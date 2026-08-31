"""The daily sales email: aggregation, the delivered filter, and the sheet."""

from __future__ import annotations

import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from decimal import Decimal as D
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook
from sqlalchemy.dialects import postgresql

from app.services.pos import daily_sales_email as dse

# ── Channel → column ─────────────────────────────────────────────────────────


def test_each_marketplace_maps_to_its_own_fixed_column():
    assert dse._column_for("aggregator", "Talabat") == "talabat"
    assert dse._column_for("aggregator", "Keeta 2.0") == "keeta"
    assert dse._column_for("aggregator", "Noon Food") == "noon_food"
    assert dse._column_for("aggregator", "Careem") == "careem"
    assert dse._column_for("aggregator", "Deliveroo") == "deliveroo"


def test_website_and_counter_are_their_own_columns():
    assert dse._column_for("online", None) == "website"
    assert dse._column_for("cashier", None) == "counter"


def test_an_unmapped_marketplace_keeps_its_name_rather_than_vanishing():
    # A new aggregator nobody has mapped is real money — it gets a column, not a
    # silent drop into nothing.
    assert dse._column_for("aggregator", "Brand New Co") == "Brand New Co"


# ── The delivered-only filter ────────────────────────────────────────────────


def test_the_filter_is_delivered_trade_only():
    sql = str(
        dse._DELIVERED.compile(
            dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
        )
    )
    # A counter check closed, or a marketplace/website order delivered.
    assert "'cashier'" in sql and "'closed'" in sql
    assert "'delivered'" in sql
    assert "'aggregator'" in sql and "'online'" in sql


def test_the_fetch_query_compiles_to_postgres():
    import asyncio

    class _Cap:
        async def execute(self, stmt):
            str(stmt.compile(dialect=postgresql.dialect()))

            class _R:
                def all(self_inner):
                    return []

            return _R()

    asyncio.run(dse._fetch(_Cap(), "2026-08-24", "2026-08-24"))


# ── build() ──────────────────────────────────────────────────────────────────

_SHJ = uuid.uuid4()
_BAR = uuid.uuid4()


def _stub_db(rows):
    branches = [
        SimpleNamespace(id=_SHJ, name="Sharjah Kitchen"),
        SimpleNamespace(id=_BAR, name="Barsha Heights"),
    ]

    class _Res:
        def __init__(self, rows=None, scal=None):
            self._rows, self._scal = rows, scal

        def all(self):
            return self._rows

        def scalars(self):
            outer = self

            class _S:
                def all(self_inner):
                    return outer._scal

            return _S()

    class _DB:
        def __init__(self):
            self.n = 0

        async def execute(self, stmt):
            self.n += 1
            return _Res(rows=rows) if self.n == 1 else _Res(scal=branches)

    return _DB()


def _row(branch, source, channel, cnt, rev, disc, agg, pay, courier, refund, mkt="0"):
    return (
        "2026-08-24",
        branch,
        source,
        channel,
        cnt,
        D(rev),
        D(disc),
        D(agg),
        D(pay),
        D(mkt),  # marketing_fee, between payment fee and courier cost
        D(courier),
        D(refund),
    )


@pytest.mark.asyncio
async def test_build_lays_out_fixed_columns_and_zero_fills():
    rows = [
        _row(_SHJ, "aggregator", "Talabat", 44, "2680", "0", "800", "100", "0", "0"),
        _row(_SHJ, "online", None, 11, "750.50", "254.50", "0", "30", "391.59", "0"),
        _row(_SHJ, "cashier", None, 1, "35", "0", "0", "0", "0", "0"),
        _row(_BAR, "aggregator", "Careem", 2, "210", "0", "55", "4", "0", "0"),
    ]
    report = await dse.build(
        _stub_db(rows), date_from="2026-08-24", date_to="2026-08-24"
    )

    assert report.columns == [
        "keeta",
        "noon_food",
        "talabat",
        "careem",
        "deliveroo",
        "website",
        "counter",
    ]
    # Both active branches appear, ordered by name; Barsha's untouched columns
    # are zero, not missing.
    by_branch = {r.branch_name: r for r in report.rows}
    assert set(by_branch) == {"Sharjah Kitchen", "Barsha Heights"}
    assert by_branch["Barsha Heights"].cells["careem"].count == 2
    assert by_branch["Barsha Heights"].cells["talabat"].revenue == 0
    assert by_branch["Sharjah Kitchen"].cells["counter"].revenue == D("35")


@pytest.mark.asyncio
async def test_net_is_revenue_less_charges_less_refunds_and_charges_sum_all_three():
    rows = [_row(_SHJ, "online", None, 11, "750.50", "0", "0", "30", "391.59", "10")]
    report = await dse.build(
        _stub_db(rows), date_from="2026-08-24", date_to="2026-08-24"
    )
    cell = {r.branch_name: r for r in report.rows}["Sharjah Kitchen"].cells["website"]
    # charges = aggregator_fee + payment_fee + courier_cost
    assert cell.charges == D("421.59")
    # net = revenue - charges - refunds
    assert cell.net == D("750.50") - D("421.59") - D("10")


@pytest.mark.asyncio
async def test_xlsx_has_the_five_sections_with_discount_and_charges_negative():
    import io

    rows = [
        _row(_SHJ, "online", None, 11, "750.50", "254.50", "0", "30", "391.59", "0"),
        _row(_SHJ, "aggregator", "Talabat", 44, "2680", "0", "800", "100", "0", "0"),
    ]
    report = await dse.build(
        _stub_db(rows), date_from="2026-08-24", date_to="2026-08-24"
    )
    wb = load_workbook(io.BytesIO(dse.to_xlsx(report)))
    ws = wb.active
    titles = [
        row[0]
        for row in ws.iter_rows(values_only=True)
        if row[0] in {s[0] for s in dse._SECTIONS}
    ]
    assert titles == [
        "Sales Revenue",
        "Order Count",
        "Sales Discount",
        "Charges",
        "Net Revenue",
    ]

    # Find the Sharjah website discount cell under the Sales Discount block.
    values = list(ws.iter_rows(values_only=True))
    disc_idx = next(i for i, r in enumerate(values) if r[0] == "Sales Discount")
    header = values[disc_idx + 1]
    website_col = header.index("website")
    shj = next(r for r in values[disc_idx:] if r[1] == "Sharjah Kitchen")
    assert shj[website_col] == -254.5  # discount shown negative

    charges_idx = next(i for i, r in enumerate(values) if r[0] == "Charges")
    shj_charges = next(r for r in values[charges_idx:] if r[1] == "Sharjah Kitchen")
    assert shj_charges[header.index("talabat")] == -900.0  # 800 + 100, negative


async def test_summary_sheet_carries_row_column_and_grand_totals():
    import io

    rows = [
        _row(_SHJ, "aggregator", "Talabat", 2, "100", "0", "10", "0", "0", "0"),
        _row(_SHJ, "online", None, 1, "50", "0", "0", "0", "5", "0"),
    ]
    report = await dse.build(
        _stub_db(rows), date_from="2026-08-24", date_to="2026-08-24"
    )
    wb = load_workbook(io.BytesIO(dse.to_xlsx(report)))
    values = list(wb["Summary"].iter_rows(values_only=True))
    rev_idx = next(i for i, r in enumerate(values) if r[0] == "Sales Revenue")
    header = values[rev_idx + 1]
    assert header[-1] == "TOTAL"  # a total column was added
    shj = next(r for r in values[rev_idx:] if r[1] == "Sharjah Kitchen")
    assert shj[-1] == 150.0  # row total across channels (100 talabat + 50 website)
    totals = next(r for r in values[rev_idx:] if r[1] == "TOTAL")
    assert totals[-1] == 150.0  # grand total


async def test_xlsx_has_the_four_detail_tabs_scoped_to_the_date():
    import io

    report = await dse.build(_stub_db([]), date_from="2026-08-27", date_to="2026-08-27")
    detail = dse.ReportDetail(
        orders=[
            # positional row, as build_detail's SELECT returns
            (
                "2026-08-27",
                "AGG-1",
                "DSO",
                "aggregator",
                "Talabat",
                "delivered",
                "closed",
                "Ali",
                D("50"),
                D("2.4"),
                D("5"),
                D("1"),
                D("0"),
                D("0"),
                "TLB-REF-9",  # external_reference (o[14])
                "9",  # aggregator_display_code (o[15])
                D("0"),  # marketing_fee (o[16])
            ),
        ],
        statements=[
            SimpleNamespace(
                channel="noon",
                statement_id="S1",
                period_start="2026-08-27",
                period_end="2026-08-27",
                gross_sales=D("100"),
                total_fees=D("10"),
                total_vat=D("5"),
                net_payable=D("85"),
                payout_transfer_id="P1",
                invoice_object_key="k",
            )
        ],
        statement_lines=[
            SimpleNamespace(
                channel="noon",
                statement_id="S1",
                external_order_id="E1",
                line_date="2026-08-27",
                line_type="sale",
                fee_category="commission",
                description="x",
                amount=D("10"),
            )
        ],
        payouts=[
            SimpleNamespace(
                channel="noon",
                transfer_id="P1",
                transfer_date="2026-08-27",
                transfer_amount=D("85"),
                transfer_status="paid",
                statement_id="S1",
                payment_reference="ref",
            )
        ],
    )
    wb = load_workbook(io.BytesIO(dse.to_xlsx(report, detail)))
    assert wb.sheetnames == [
        "Summary",
        "Orders",
        "Statements",
        "Statement Lines",
        "Payouts",
    ]
    assert wb["Orders"].cell(row=2, column=2).value == "AGG-1"
    assert wb["Orders"].cell(row=2, column=3).value == "TLB-REF-9"  # aggregator ref
    assert wb["Orders"].cell(row=2, column=4).value == "9"  # aggregator code
    assert wb["Orders"].cell(row=2, column=6).value == "Talabat"  # channel label
    assert wb["Orders"].cell(row=2, column=13).value == 0.0  # marketing fee
    assert wb["Orders"].cell(row=2, column=16).value == 44.0  # net = 50-5-1-0
    assert wb["Statements"].cell(row=2, column=2).value == "S1"
    assert wb["Statement Lines"].cell(row=2, column=8).value == 10.0
    assert wb["Payouts"].cell(row=2, column=4).value == 85.0


# ── Scheduling: a day sends once the last branch's close + buffer has passed ──

_DUBAI = ZoneInfo("Asia/Dubai")


@asynccontextmanager
async def _lock_ok(*_a, **_k):
    yield True


def _branches_db(branches):
    class _Scalars:
        def all(self):
            return branches

    class _Res:
        def scalars(self):
            return _Scalars()

    async def execute(_stmt):
        return _Res()

    return SimpleNamespace(execute=execute)


def _branch(name, opening_from="09:00", opening_to="23:00"):
    return SimpleNamespace(name=name, opening_from=opening_from, opening_to=opening_to)


async def _run_tick(*, now: datetime, branches: list, already_sent: set[str]):
    """Drive `_tick` at a given instant with real branch hours, the send and the
    already-sent guard stubbed, and read back which business dates got mailed."""
    sent: list[str] = []

    async def fake_send(_db, *, date_from, date_to, recipients):
        sent.append(date_from)
        return {"sent": []}

    with (
        patch.object(
            dse.business_day_service,
            "resolve_timezone",
            AsyncMock(return_value=_DUBAI),
        ),
        patch.object(
            dse,
            "_already_sent",
            AsyncMock(side_effect=lambda _db, d: d in already_sent),
        ),
        patch.object(dse, "send", fake_send),
        patch.object(dse.advisory_lock, "held", _lock_ok),
    ):
        await dse._tick(_branches_db(branches), now=now)
    return sent


@pytest.mark.asyncio
async def test_it_sends_the_day_the_morning_after_at_the_send_hour():
    # Branches close 22:00 and 23:00 on the 25th. The report for the 25th is held
    # past its close+buffer until 01:00 Dubai on the 26th (the send hour, giving
    # the overnight aggregator scrape the night to land). At 01:05 on the 26th it
    # is due; the 26th itself is still in progress and held. Prior days are mailed.
    sent = await _run_tick(
        now=datetime(2026, 8, 26, 1, 5, tzinfo=_DUBAI),
        branches=[_branch("A", opening_to="22:00"), _branch("B", opening_to="23:00")],
        already_sent={"2026-08-23", "2026-08-24"},
    )
    assert sent == ["2026-08-25"]


@pytest.mark.asyncio
async def test_it_holds_today_until_the_send_hour_even_after_close():
    # 00:30 on the 26th is well past the 25th's close+buffer (23:45) but before the
    # 01:00 send hour, so the 25th is still held — the send hour, not the close, is
    # the binding floor now. Prior days already sent, so nothing goes out.
    sent = await _run_tick(
        now=datetime(2026, 8, 26, 0, 30, tzinfo=_DUBAI),
        branches=[_branch("A", opening_to="23:00")],
        already_sent={"2026-08-23", "2026-08-24"},
    )
    assert sent == []


@pytest.mark.asyncio
async def test_it_catches_up_completed_days_oldest_first_never_the_in_progress_day():
    # Midday: today's close (23:00) has not passed, so the in-progress day is held,
    # while the three completed days before it are all due and unmailed.
    sent = await _run_tick(
        now=datetime(2026, 8, 25, 12, 0, tzinfo=_DUBAI),
        branches=[_branch("A", opening_to="23:00")],
        already_sent=set(),
    )
    assert sent == ["2026-08-22", "2026-08-23", "2026-08-24"]
    assert "2026-08-25" not in sent


@pytest.mark.asyncio
async def test_it_skips_days_already_mailed():
    sent = await _run_tick(
        now=datetime(2026, 8, 25, 12, 0, tzinfo=_DUBAI),
        branches=[_branch("A", opening_to="23:00")],
        already_sent={"2026-08-22", "2026-08-23"},
    )
    assert sent == ["2026-08-24"]


@pytest.mark.asyncio
async def test_it_sends_nothing_when_the_window_is_all_mailed():
    sent = await _run_tick(
        now=datetime(2026, 8, 25, 12, 0, tzinfo=_DUBAI),
        branches=[_branch("A", opening_to="23:00")],
        already_sent={"2026-08-22", "2026-08-23", "2026-08-24"},
    )
    assert sent == []


@pytest.mark.asyncio
async def test_a_past_midnight_branch_holds_its_day_until_it_actually_closes():
    # A kitchen open 09:00 → 02:00 shuts the 24th's trading at 02:00 on the 25th,
    # later than a 23:00 branch. The 24th's report must wait for that 02:00 close
    # (+ buffer, so 02:45), not fire at the earlier branch's close.
    late = _branch("Late", opening_from="09:00", opening_to="02:00")
    early = _branch("Early", opening_to="23:00")

    # 02:30 on the 25th: past midnight but before the 02:45 due time — the 24th is
    # held, while the 23rd (whose 02:00 close was the 24th) is already due.
    held = await _run_tick(
        now=datetime(2026, 8, 25, 2, 30, tzinfo=_DUBAI),
        branches=[late, early],
        already_sent=set(),
    )
    assert "2026-08-24" not in held
    assert "2026-08-23" in held

    # 03:00 on the 25th: the 02:00 close plus buffer has passed, so the 24th sends.
    due = await _run_tick(
        now=datetime(2026, 8, 25, 3, 0, tzinfo=_DUBAI),
        branches=[late, early],
        already_sent={"2026-08-22", "2026-08-23"},
    )
    assert due == ["2026-08-24"]
