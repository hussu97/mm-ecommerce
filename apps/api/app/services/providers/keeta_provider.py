"""Keeta, the one channel a captured session cannot replay over httpx.

Keeta's merchant portal (`merchant.mykeeta.com`) is a thin skin over Meituan's
infrastructure, and Meituan guards every XHR with `mtgsig` — a per-request
signature the page's own obfuscated JS computes from the method, the path, the
body, a rolling client fingerprint and a clock, then stamps onto the request as
an `mtgsig` header. The signer runs *in the page*; there is no token to lift and
resend, because the signature is bound to the exact request it rides on and to
browser state (the WASM/JS sensor's evolving seed) that a headless `httpx` call
does not have. Replaying a captured session the way Careem or Deliveroo are
replayed gets a signed-out or risk-controlled response, never the data — the
cookie is fine, but the request that carries it is unsigned and rejected.

So Keeta breaks the shape every other provider fits. The rest of the fleet
(`careem`, `deliveroo`, `talabat`, `noon`) is fetched on the hourly httpx path
from a bootstrap-captured session; Keeta is **not registered in that loop**. Its
data is pulled *in-page* by the bootstrap worker: the same browser that holds
the live session evaluates the fetch in the page context, where the portal's own
JS signs it, and hands the raw JSON back out. That is the only place `mtgsig`
exists, so that is where the fetch has to happen.

What this module therefore is, and is not:

- It is **the parsing half** — the reusable Keeta-JSON-to-DTO translation. The
  bootstrap worker calls `parse_orders(payload)` with a `getOrders` response and
  `parse_finance(payload)` with a finance-download response it fetched in-page,
  and gets back the same channel-neutral DTOs every other provider produces, so
  everything downstream of the provider edge stays ignorant of Keeta's
  vocabulary (its minor-unit money, its `baseOrder`/`merchantOrder`/`feeDtl`
  envelope, its Meituan field spellings).
- It is **not a working httpx client**. `fetch_sales` / `fetch_finance` exist to
  satisfy `BaseAggregatorClient`, but they raise `AggregatorUnavailableError`
  with a message naming the constraint rather than pretending to fetch — a Keeta
  session on the httpx path is a misconfiguration, and this makes it loud.

Parsing is deliberately defensive and unit-explicit. Keeta's internal JSON
states money in *fils* (minor units); `feeDtl.merchantFee` is divided to AED at
the point the unit is known (`_from_minor_units`) rather than by guessing from
magnitude. Money the payload does not carry is left `None` (unknown), never `0`
(charged nothing) — the same rule the rest of the ingest keeps. Every DTO
retains its source row in `raw`, so the mapping is refined against real payloads
without re-fetching. Dates are resolved to Dubai wall-clock, because that is what
a business date means to the people reading the reconciliation.
"""

from __future__ import annotations

import base64
import json
import logging
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any
from zoneinfo import ZoneInfo

from app.models.aggregator import CHANNEL_KEETA, GRAIN_LINE
from app.services.aggregators.modifiers import expand_modifiers
from app.services.aggregators.normalized import (
    FinanceResult,
    PayoutsResult,
    SalesResult,
    StandardOrder,
    StandardOrderItem,
    StandardPayout,
    StandardStatement,
    StandardStatementLine,
    StatementsResult,
)
from app.services.aggregators.session_store import LoadedSession
from app.services.providers.aggregator_base import (
    AggregatorUnavailableError,
    BaseAggregatorClient,
)

logger = logging.getLogger(__name__)

#: The shop's clock. Every date this module emits is a Dubai wall-clock date,
#: because that is the business date the reports are read against.
_BUSINESS_TZ = ZoneInfo("Asia/Dubai")

_UNAVAILABLE = (
    "Keeta is ingested in-page by the bootstrap worker "
    "(mtgsig request signing cannot be replayed over httpx)"
)

# ── Field spellings ─────────────────────────────────────────────────────────
# Keeta/Meituan spell the same field many ways across surfaces; keys are matched
# case- and separator-insensitively (`_canonical_key`), so `orderViewId`,
# `order_view_id` and `ORDERVIEWID` all hit.
_ORDER_ID_KEYS = (
    "orderId",
    "order_id",
    "orderNo",
    "order_no",
    "orderNumber",
    "order_number",
    "orderViewId",
    "orderViewIdStr",
    "wmOrderIdView",
    "id",
)
_ORDER_DATE_KEYS = (
    "businessDate",
    "business_date",
    "completeDate",
    "completedDate",
    "completedTime",
    "orderDate",
    "order_date",
    "orderTime",
    "order_time",
    "orderCreateTime",
    "createTime",
    "ctime",
    "createdAt",
    "completedStatusTime",
    "placedAt",
)
_OUTLET_ID_KEYS = (
    "shopId",
    "shop_id",
    "storeId",
    "store_id",
    "merchantId",
    "merchant_id",
)
_STATEMENT_ID_KEYS = (
    "statementId",
    "statement_id",
    "billId",
    "bill_id",
    "invoiceId",
    "invoice_id",
    "settleId",
)
_TRANSFER_ID_KEYS = (
    "transferId",
    "transfer_id",
    "paymentId",
    "payment_id",
    "payoutId",
    "payout_id",
    "settleId",
)
_CURRENCY_KEYS = ("currency", "currencyCode")
_GROSS_KEYS = (
    "grossSales",
    "gross_sales",
    "originAmount",
    "originalAmount",
    "orderAmount",
    "order_amount",
    "totalAmount",
    "total_amount",
    "foodAmount",
    "actualPayAmount",
    "productPrice",
    "productPriceSubTotal",
)


# ── Scalar coercion (all money is Decimal | None; None means "not stated") ───
def _canonical_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _get_value(row: dict[str, Any], key: str) -> Any:
    """The value under `key`, matched case/separator-insensitively."""
    canonical = _canonical_key(key)
    for current_key, value in row.items():
        if _canonical_key(current_key) == canonical:
            return value
    return None


def _money(value: Any) -> Decimal | None:
    """A money amount as Decimal, or None for anything not a clean number.

    Accepts the numbers already-parsed by the flatteners (int/float/Decimal) and
    the human strings the portal sometimes carries (`"1,234.50 AED"`, `"(5.00)"`
    for a credit). None is the honest answer for a missing or unreadable value —
    the caller must not read it as zero.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    cleaned = (
        text.replace("AED", "")
        .replace("د.إ", "")
        .replace(",", "")
        .replace("(", "-")
        .replace(")", "")
        .strip()
    )
    cleaned = re.sub(r"[^\d.\-]", "", cleaned)
    if cleaned in {"", "-", ".", "-."}:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _first_money(row: dict[str, Any], keys: tuple[str, ...]) -> Decimal | None:
    """The first key present as a readable number, else None (unknown)."""
    for key in keys:
        value = _get_value(row, key)
        if value is None or str(value).strip() == "":
            continue
        parsed = _money(value)
        if parsed is not None:
            return parsed
    return None


def _abs_money(value: Decimal | None) -> Decimal | None:
    """Fees/discounts are carried as magnitudes; None stays unknown."""
    return abs(value) if value is not None else None


def _from_minor_units(value: Any) -> Decimal | None:
    """Convert an amount Keeta states in fils to AED, at a known-unit call site."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return Decimal(str(value)) / Decimal(100)


def _first_text(row: dict[str, Any], keys: tuple[str, ...]) -> str | None:
    for key in keys:
        value = _get_value(row, key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return None


def _first_datetime(row: dict[str, Any], keys: tuple[str, ...]) -> datetime | None:
    for key in keys:
        parsed = _parse_datetime(_get_value(row, key))
        if parsed:
            return parsed
    return None


def _first_date(row: dict[str, Any], keys: tuple[str, ...]) -> date | None:
    for key in keys:
        parsed = _parse_datetime(_get_value(row, key))
        if parsed:
            return parsed.date()
    return None


def _as_business_local(parsed: datetime) -> datetime:
    """A datetime as naive Dubai wall-clock time.

    An offset-aware instant is converted (not truncated) into Dubai time, so a
    late-night order carrying a `Z` offset books to the right business date;
    naive input is assumed already-local, which is what an offset-less string
    from the portal means.
    """
    if parsed.tzinfo is None:
        return parsed
    return parsed.astimezone(_BUSINESS_TZ).replace(tzinfo=None)


def _parse_datetime(value: Any) -> datetime | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return _as_business_local(value)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric > 10_000_000_000:  # milliseconds, not seconds
            numeric = numeric / 1000
        try:
            # An epoch is an instant; turning it into a wall clock needs a zone,
            # and it must be Dubai's — never the host machine's — so the same
            # order never dates differently on a laptop and a UTC server.
            return datetime.fromtimestamp(numeric, _BUSINESS_TZ).replace(tzinfo=None)
        except (OSError, ValueError, OverflowError):
            return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("Z", "+00:00")
    for fmt in (
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return _as_business_local(datetime.strptime(text, fmt))
        except ValueError:
            continue
    try:
        return _as_business_local(datetime.fromisoformat(text))
    except ValueError:
        return None


def _date_str(value: date | None) -> str | None:
    return value.strftime("%Y-%m-%d") if value else None


# ── Finance file-payload parsing (bill xlsx + commission zip) ────────────────
# The bootstrap worker downloads Keeta's presigned billing-report XLSX and
# commission-invoice ZIP (both plain-httpx once the LIST call yields the signed
# URL) and pushes them base64-encoded. These helpers turn the worker's date
# spellings into `date` objects and periods; the money inside the XLSX is stated
# in AED **major units** (the sheet shows "40.0", "-9.0", "26.2"), so it is read
# with `_money` directly — never `_from_minor_units`, which is only for the fils
# in the getOrders path.
def _keeta_short_date(value: Any) -> date | None:
    """Parse Keeta's report date — `'15 Aug 2026'`, optionally `'... at HH:MM:SS'`."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    head = text.split(" at ")[0].strip()
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(head, fmt).date()
        except ValueError:
            continue
    return None


def _period_from_display(value: Any) -> tuple[date | None, date | None]:
    """`'15 Aug 2026 ~ 22 Aug 2026'` → (start, end); `(None, None)` if unparsable."""
    if value is None:
        return (None, None)
    text = str(value).strip()
    for sep in ("~", " - ", " to "):
        if sep in text:
            left, right = text.split(sep, 1)
            return (_keeta_short_date(left), _keeta_short_date(right))
    return (None, None)


def _dotted_date(value: Any) -> date | None:
    """Parse Keeta's billing-cycle date — `'2026.08.15'` → date, else None."""
    if value is None:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y.%m.%d").date()
    except ValueError:
        return None


def _billing_cycle_bounds(value: Any) -> tuple[date | None, date | None]:
    """`'2026.08.15~2026.08.21'` → (start, end); `(None, None)` if unparsable.

    This is the "Invoice Details" `Billing Cycle` spelling — dotted `YYYY.MM.DD`
    around a `~` — distinct from `displayTimeText`'s `'15 Aug 2026 ~ 22 Aug 2026'`
    that `_period_from_display` handles.
    """
    if value is None:
        return (None, None)
    text = str(value).strip()
    if "~" not in text:
        return (None, None)
    left, right = text.split("~", 1)
    return (_dotted_date(left), _dotted_date(right))


def _period_from_yyyymm(value: Any) -> tuple[date | None, date | None]:
    """A `YYYYMM` int/str (e.g. `202607`) → (first day, last day of that month)."""
    try:
        n = int(str(value).strip())
    except (TypeError, ValueError):
        return (None, None)
    year, month = divmod(n, 100)
    if not (1 <= month <= 12 and 2000 <= year <= 2100):
        return (None, None)
    start = date(year, month, 1)
    first_of_next = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return (start, first_of_next - timedelta(days=1))


#: Finance file payloads carry the report/invoice bytes; their presence is what
#: routes `parse_finance` down the file path rather than the row-tree path.
_FINANCE_FILE_KEYS = ("bill_xlsx_b64", "invoice_zip_b64", "invoice_pdf_b64")


def _has_finance_file(row: dict[str, Any]) -> bool:
    return any(_get_value(row, key) for key in _FINANCE_FILE_KEYS)


def _strip_blobs(row: dict[str, Any]) -> dict[str, Any]:
    """The row minus the base64 file blobs — so `raw` JSONB stays small."""
    if not isinstance(row, dict):
        return row
    return {k: v for k, v in row.items() if k not in _FINANCE_FILE_KEYS}


#: "Order Summary" sheet: header on row 3, data from row 4. Money columns are
#: 1-indexed (matching the reverse-engineered layout). Each becomes its own
#: statement line so gross / commission / bank-fee / net stay distinguishable.
_BILL_SHEET = "Order Summary"
_BILL_ORDER_NO_COL = 9  # 16-digit Order Number == sales external_order_id
_BILL_TXN_DATE_COL = 6  # "15 Aug 2026"
#: category → (source column, line_type, fee_category).
_BILL_LINE_SPEC: tuple[tuple[str, int, str, str], ...] = (
    ("gross", 16, "sales", "gross_sales"),  # Original item price (VAT incl)
    ("commission", 35, "commission", "commission"),  # Total Commission (VAT incl)
    ("bank_fee", 22, "fee", "bank_fee"),  # Bank fee (VAT incl)
    ("net", 33, "payout", "net_payable"),  # Payable to Restaurant (net)
)


def _cell(values: tuple[Any, ...], col: int) -> Any:
    """A 1-indexed column from an openpyxl `values_only` row, padded to None."""
    idx = col - 1
    return values[idx] if 0 <= idx < len(values) else None


def _parse_bill_xlsx(
    xlsx_bytes: bytes, statement_id: str
) -> list[StandardStatementLine]:
    """Parse the "Order Summary" sheet into per-order statement lines.

    One order yields up to four lines — gross, commission, bank fee, net —
    keyed to the 16-digit Order Number (the join key to the sales
    `external_order_id`). Blank money cells are left out (a missing figure is
    unknown, never 0); a real 0.0 in the sheet is kept as stated. Amounts are AED
    major units, so `_money` reads them without any /100 conversion.
    """
    import io

    from openpyxl import load_workbook

    # NOT read_only: Keeta's file omits the `<dimension>` tag, so read-only mode
    # reports a 1×1 sheet and reads nothing. The files are small (~16 KB), so a
    # full parse is cheap and correct.
    try:
        workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception:  # noqa: BLE001 — a corrupt sheet must not abort the parse
        logger.warning("keeta: unreadable bill xlsx for %s", statement_id)
        return []
    try:
        if _BILL_SHEET not in workbook.sheetnames:
            logger.warning(
                "keeta: bill xlsx for %s has no %r sheet", statement_id, _BILL_SHEET
            )
            return []
        sheet = workbook[_BILL_SHEET]
        lines: list[StandardStatementLine] = []
        for values in sheet.iter_rows(min_row=4, values_only=True):
            order_no = _cell(values, _BILL_ORDER_NO_COL)
            if order_no is None or not str(order_no).strip():
                continue
            order_no = str(order_no).strip()
            line_date = _date_str(_keeta_short_date(_cell(values, _BILL_TXN_DATE_COL)))
            for category, col, line_type, fee_category in _BILL_LINE_SPEC:
                amount = _money(_cell(values, col))
                if amount is None:
                    continue
                lines.append(
                    StandardStatementLine(
                        source_key=f"{statement_id}:{order_no}:{category}",
                        statement_id=statement_id,
                        external_order_id=order_no,
                        line_date=line_date,
                        line_type=line_type,
                        fee_category=fee_category,
                        amount=amount,
                        currency="AED",
                    )
                )
        return lines
    finally:
        workbook.close()


# "Invoice Details" sheet: the weekly *settlement* view — one row per shop per
# transaction day, carrying the net "Payable to Restaurant" and the billing cycle
# it settles under. Header on row 1, data from row 2; columns are 1-indexed to
# match the real file's layout (confirmed against the downloaded bill.xlsx). This
# is the payout side of the bill — the TOTAL Keeta actually transfers per weekly
# cycle, which the "Order Summary" per-order parser never sums.
_PAYOUT_SHEET = "Invoice Details"
_PAYOUT_RESTAURANT_ID_COL = 4  # "Restaurant ID" — shop id fallback
_PAYOUT_PAYABLE_COL = 8  # "Payable to Restaurant" (net, AED major units)
_PAYOUT_STATUS_COL = 9  # "Settled" / "Settlement pending"
_PAYOUT_CYCLE_COL = 10  # "2026.08.15~2026.08.21"


def _parse_bill_payouts(
    xlsx_bytes: bytes,
    statement_id: str,
    payload_shop_id: str | None,
    task_view_id: str | None,
) -> list[StandardPayout]:
    """Parse the "Invoice Details" sheet into one payout per weekly billing cycle.

    Keeta settles weekly: every billing cycle in the sheet is a distinct transfer
    to the restaurant, so rows are grouped by their `Billing Cycle` value. A
    group's `transfer_amount` is the sum of its "Payable to Restaurant" figures
    (AED **major** units — read with `_money`, never `_from_minor_units`), its
    `transfer_date`/`payment_due_date` is the cycle-end date, and its
    `transfer_status` is `"settled"` only when every row of the cycle reads
    `"Settled"` (any pending row makes the whole cycle `"pending"`).

    The `transfer_id` is `KEETA_BILL_{shopId}_{cycleEnd}` — stable and unique per
    (shop, week), so re-ingesting the same bill upserts in place rather than
    duplicating. Each payout carries the caller's `statement_id`, coupling it to
    the weekly statement, and `payment_reference` is the download task's
    `taskViewId`. A cycle whose payables are all blank stays `None` (unknown),
    never a fabricated 0; a cycle whose end date will not parse is skipped (it
    cannot be keyed) rather than emitted under a guessed id.
    """
    import io

    from openpyxl import load_workbook

    # NOT read_only — same `<dimension>`-tag omission as the Order Summary parser.
    try:
        workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception:  # noqa: BLE001 — a corrupt sheet yields no payouts, not a crash
        logger.warning("keeta: unreadable bill xlsx (payouts) for %s", statement_id)
        return []
    try:
        if _PAYOUT_SHEET not in workbook.sheetnames:
            return []
        sheet = workbook[_PAYOUT_SHEET]
        # cycle string → aggregation state, ordered by first appearance.
        groups: dict[str, dict[str, Any]] = {}
        for values in sheet.iter_rows(min_row=2, values_only=True):
            cycle_cell = _cell(values, _PAYOUT_CYCLE_COL)
            cycle_key = str(cycle_cell).strip() if cycle_cell is not None else ""
            if not cycle_key:
                continue
            group = groups.setdefault(
                cycle_key,
                {"total": None, "all_settled": True, "shop_id": None},
            )
            payable = _money(_cell(values, _PAYOUT_PAYABLE_COL))
            if payable is not None:
                group["total"] = (
                    payable if group["total"] is None else group["total"] + payable
                )
            status_cell = _cell(values, _PAYOUT_STATUS_COL)
            status = _normalize_status(
                str(status_cell).strip() if status_cell is not None else None
            )
            if status != "settled":
                group["all_settled"] = False
            if group["shop_id"] is None:
                rid = _cell(values, _PAYOUT_RESTAURANT_ID_COL)
                if rid is not None and str(rid).strip():
                    group["shop_id"] = str(rid).strip()

        payouts: list[StandardPayout] = []
        for cycle_key, group in groups.items():
            _start, cycle_end = _billing_cycle_bounds(cycle_key)
            cycle_end_str = _date_str(cycle_end)
            if cycle_end_str is None:
                logger.warning(
                    "keeta: bill payout cycle %r unparsable for %s — skipped",
                    cycle_key,
                    statement_id,
                )
                continue
            shop_id = payload_shop_id or group["shop_id"] or "unknown"
            payouts.append(
                StandardPayout(
                    transfer_id=f"KEETA_BILL_{shop_id}_{cycle_end_str}",
                    statement_id=statement_id,
                    transfer_date=cycle_end_str,
                    payment_due_date=cycle_end_str,
                    transfer_amount=group["total"],
                    transfer_status="settled" if group["all_settled"] else "pending",
                    payment_reference=task_view_id or statement_id,
                    currency="AED",
                )
            )
        return payouts
    finally:
        workbook.close()


def _normalize_status(value: str | None) -> str | None:
    if not value:
        return None
    normalized = re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")
    return normalized or None


# Keeta/Meituan numeric order-status codes → readable lowercase names.
# Justified from orders_sample.json: `merchantOrderTraces` records the lifecycle
# 10 → 20 → 30 for a live order, with `unconfirmedStatusTime` stamping entry to
# status 20 and `confirmedStatusTime` stamping entry to status 30; the settled
# history rows all carry status 40. Codes not evidenced by the sample (a
# cancellation/refund code among them) are intentionally left unmapped and fall
# back to the raw numeric string rather than being guessed.
_STATUS_CODES = {
    "10": "submitted",
    "20": "pending",
    "30": "confirmed",
    "40": "completed",
}


def _decode_status(value: str | None) -> str | None:
    """Decode Keeta's numeric status to a readable name, else keep the raw value.

    A non-numeric status (an already-worded fixture, a future spelling) passes
    through `_normalize_status` unchanged; an unknown numeric code is returned as
    its normalized digits rather than a made-up name.
    """
    normalized = _normalize_status(value)
    if normalized is None:
        return None
    return _STATUS_CODES.get(normalized, normalized)


def _mods_text(item: dict[str, Any]) -> str | None:
    """Raw option dump for debugging — real lines carry `groups`/`spuPvList`,
    older/synthetic fixtures `modifiers`/`attributes`; serialize whichever is
    present."""
    for key in ("groups", "spuPvList", "modifiers", "attributes"):
        value = _get_value(item, key)
        if value not in (None, "", []):
            return json.dumps(value, ensure_ascii=False, default=str)
    return None


def _item_modifier_sources(item: dict[str, Any]) -> list[Any]:
    """Collect the raw chosen-option payloads on one Keeta product line.

    Real Keeta lines carry chosen options in two places, both verified present on
    every `products[]` entry in orders_sample.json (though empty in that sample,
    so the per-option leaf spelling could not be pinned to it):
      - `groups[]`  — add-on / combo groups; each group wraps its option items in
        a nested list, so we flatten one level to the leaf options.
      - `spuPvList[]` — SPU property/variant selections, already option-shaped.
    Older/synthetic fixtures use `modifiers`/`attributes`; those are kept as
    fallbacks. Every source is funnelled through `expand_modifiers`, so the leaf
    option shape is matched with the same name/qty/price/ref fallbacks the other
    channels use (an unrecognised leaf is dropped, never invented).
    """
    sources: list[Any] = []

    groups = _get_value(item, "groups")
    if isinstance(groups, list):
        for group in groups:
            if not isinstance(group, dict):
                continue
            nested: list[Any] | None = None
            for key in ("foods", "spus", "skus", "products", "items", "options"):
                value = _get_value(group, key)
                if isinstance(value, list) and value:
                    nested = value
                    break
            # A group with a nested option list contributes its leaves; a flat
            # group dict is itself treated as the chosen option.
            sources.extend(nested if nested is not None else [group])

    spu_pv = _get_value(item, "spuPvList")
    if isinstance(spu_pv, list):
        sources.extend(pv for pv in spu_pv if isinstance(pv, dict))

    legacy = _get_value(item, "modifiers") or _get_value(item, "attributes")
    if legacy not in (None, "", []):
        sources.extend(legacy if isinstance(legacy, list) else [legacy])

    return sources


def _customer_name(row: dict[str, Any]) -> str | None:
    """Recipient/user display name from the getOrders envelope.

    `recipientInfo.name` is the delivery recipient (sometimes a portal code like
    `J4P773781744`, sometimes masked to `***`); `userInfo.userName` is the
    account name (usually masked). We store whatever is present — a masked value
    is what the portal gave, not an invented one.
    """
    recipient = row.get("recipientInfo")
    if isinstance(recipient, dict):
        name = _first_text(recipient, ("name",))
        if name:
            return name
    user_info = row.get("userInfo")
    if isinstance(user_info, dict):
        return _first_text(user_info, ("userName", "user_name", "name"))
    return None


def _customer_phone(row: dict[str, Any]) -> str | None:
    """Best-available customer phone from the getOrders envelope.

    Priority: the masked privacy line `openPrivacyNumber` when the order carries
    one (0/absent means none), then the recipient's number (prefixed with its
    `interCode` when the number is not itself masked), then the account's masked
    number. None means the payload stated nothing.
    """
    open_privacy = _get_value(row, "openPrivacyNumber")
    if open_privacy not in (None, 0, "0", ""):
        text = str(open_privacy).strip()
        if text:
            return text
    for source_key in ("recipientInfo", "userInfo"):
        source = row.get(source_key)
        if not isinstance(source, dict):
            continue
        phone = _first_text(source, ("privacyPhone", "phone", "userPhone"))
        if not phone:
            continue
        inter_code = _first_text(source, ("interCode", "inter_code"))
        if (
            inter_code
            and "*" not in phone
            and not phone.startswith("+")
            and not phone.startswith(inter_code)
        ):
            return f"+{inter_code}{phone}"
        return phone
    return None


# ── Row extraction (the getOrders / finance envelope, flattened) ─────────────
def _iter_row_candidates(payload: Any) -> list[dict[str, Any]]:
    """Every dict anywhere in the payload tree — the portal buries the order
    list at varying depths (`data.list[...]`), so the walk is unconditional."""
    rows: list[dict[str, Any]] = []
    if isinstance(payload, dict):
        rows.append(payload)
        for value in payload.values():
            rows.extend(_iter_row_candidates(value))
    elif isinstance(payload, list):
        for value in payload:
            rows.extend(_iter_row_candidates(value))
    return rows


def _normalize_product_item(item: dict[str, Any]) -> dict[str, Any]:
    """One `products[]` line, with its Meituan spellings aliased and its
    minor-unit prices divided to AED."""
    normalized = dict(item)
    if "count" in item and "quantity" not in normalized:
        normalized["quantity"] = item["count"]
    if "name" in item and "itemName" not in normalized:
        normalized["itemName"] = item["name"]
    if "skuId" in item and "itemId" not in normalized:
        normalized["itemId"] = item["skuId"]
    for key in ("price", "originPrice"):
        converted = _from_minor_units(item.get(key))
        if converted is not None:
            normalized[key] = converted
    price_with_group = item.get("priceWithGroup")
    if isinstance(price_with_group, dict):
        amount = _from_minor_units(price_with_group.get("amount"))
        unit_price = _from_minor_units(price_with_group.get("unitPrice"))
        if amount is not None:
            normalized["totalAmount"] = amount
        if unit_price is not None:
            normalized["unitPrice"] = unit_price
    return normalized


def _flatten_keeta_order(row: dict[str, Any]) -> dict[str, Any] | None:
    """Collapse Keeta's `baseOrder` + `merchantOrder` + `feeDtl.merchantFee`
    envelope into one flat row, dividing the fee block's fils to AED.

    Returns None for a row that is not this envelope, so the caller can fall
    through to treating an already-flat row as a candidate.
    """
    base_order = row.get("baseOrder")
    merchant_order = row.get("merchantOrder")
    if not isinstance(base_order, dict) or not isinstance(merchant_order, dict):
        return None
    flattened: dict[str, Any] = {
        **base_order,
        **merchant_order,
        "items": [
            _normalize_product_item(item)
            for item in row.get("products", []) or []
            if isinstance(item, dict)
        ],
    }
    # Customer identity lives on the envelope, not in base/merchant order — carry
    # it through so `_order_from` can read recipient/user name and phone.
    for passthrough in ("recipientInfo", "userInfo"):
        value = row.get(passthrough)
        if isinstance(value, dict):
            flattened[passthrough] = value
    fee_detail = row.get("feeDtl")
    merchant_fee = (
        fee_detail.get("merchantFee") if isinstance(fee_detail, dict) else None
    )
    if isinstance(merchant_fee, dict):
        for field_name in (
            "productPrice",
            "commission",
            "brokerage",
            "activityFee",
            "bankTransactionFee",
            "platformServiceFee",
            "earnings",
            "total",
        ):
            value = merchant_fee.get(field_name)
            converted = _from_minor_units(value)
            if converted is not None:
                flattened[field_name] = converted
            elif value is not None:
                flattened[field_name] = value
    return flattened


def _looks_like_row(row: dict[str, Any]) -> bool:
    """Whether an already-flat dict is itself an order/statement/transfer row —
    i.e. carries one of the id keys, or a shop id plus an amount."""
    keys = {_canonical_key(key) for key in row}
    id_keys = {
        _canonical_key(k)
        for k in _ORDER_ID_KEYS + _STATEMENT_ID_KEYS + _TRANSFER_ID_KEYS
    }
    return bool(
        keys.intersection(id_keys)
        or (
            keys.intersection({"shopid", "storeid", "shopname", "storename"})
            and keys.intersection(
                {"amount", "totalamount", "settleamount", "paymentamount"}
            )
        )
    )


def _extract_rows(payload: Any) -> list[dict[str, Any]]:
    """The order/statement rows a Keeta response carries, envelope-flattened.

    Walks the whole tree because the list lives at a surface-dependent depth,
    flattening every `baseOrder`/`merchantOrder` envelope it meets and also
    keeping any already-flat row that looks like a record.
    """
    rows: list[dict[str, Any]] = []
    for candidate in _iter_row_candidates(payload):
        flattened = _flatten_keeta_order(candidate)
        if flattened:
            rows.append(flattened)
        if _looks_like_row(candidate):
            rows.append(candidate)
    return rows


def _nested_items(row: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ("items", "itemList", "skuList", "products", "dishes"):
        value = _get_value(row, key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def _archive_keeta_invoice(row: dict[str, Any], statement_id: str) -> "Any | None":
    """Archive a Keeta commission invoice and return a `StoredStatementInvoice`.

    Priority order:
    1. PDF bytes from `invoice_pdf_b64` (base64 encoded by the bootstrap worker
       after downloading the commission invoice in-page).
    2. ZIP bytes from `invoice_zip_b64` as primary when no PDF is available.
    3. JSON of the bill-list row as a weak audit fallback — the bootstrap did not
       resolve the PDF yet, but at least finance has the metadata.

    Archive failures are logged and return None — a missing R2 document must
    not abort the finance parse.
    """
    from app.services.aggregators.statement_docs import store_statement_invoice

    pdf_bytes: bytes | None = None
    zip_bytes: bytes | None = None

    pdf_b64 = _get_value(row, "invoice_pdf_b64")
    if pdf_b64:
        try:
            pdf_bytes = base64.b64decode(pdf_b64)
        except Exception:
            logger.warning(
                "keeta: invalid invoice_pdf_b64 for %s — skipping PDF", statement_id
            )

    zip_b64 = _get_value(row, "invoice_zip_b64")
    if zip_b64:
        try:
            zip_bytes = base64.b64decode(zip_b64)
        except Exception:
            logger.warning(
                "keeta: invalid invoice_zip_b64 for %s — skipping zip", statement_id
            )

    if pdf_bytes:
        extra = [("invoice.zip", zip_bytes, "application/zip")] if zip_bytes else None
        try:
            return store_statement_invoice(
                channel=CHANNEL_KEETA,
                statement_id=statement_id,
                filename=f"{statement_id}.pdf",
                body=pdf_bytes,
                content_type="application/pdf",
                extra_files=extra,
            )
        except Exception:
            logger.warning(
                "keeta: PDF archival failed for %s — trying JSON fallback", statement_id
            )

    if zip_bytes and not pdf_bytes:
        try:
            return store_statement_invoice(
                channel=CHANNEL_KEETA,
                statement_id=statement_id,
                filename=f"{statement_id}.zip",
                body=zip_bytes,
                content_type="application/zip",
            )
        except Exception:
            logger.warning(
                "keeta: zip archival failed for %s — trying JSON fallback", statement_id
            )

    # JSON fallback — the PDF is not yet downloaded; archive the bill-list row so
    # finance has at least the metadata as an auditable document.
    try:
        json_bytes = json.dumps(row, default=str, ensure_ascii=False).encode("utf-8")
        return store_statement_invoice(
            channel=CHANNEL_KEETA,
            statement_id=statement_id,
            filename=f"{statement_id}.json",
            body=json_bytes,
            content_type="application/json",
        )
    except Exception:
        logger.warning("keeta: JSON archival fallback failed for %s", statement_id)
        return None


class KeetaClient(BaseAggregatorClient):
    """Keeta's parser, and a fetch path that refuses rather than lies.

    `parse_orders` / `parse_finance` are the surface the bootstrap worker calls
    with JSON it fetched in-page (where `mtgsig` is signed). `fetch_sales` /
    `fetch_finance` exist only to complete `BaseAggregatorClient`; both raise,
    because Keeta is never on the httpx ingest loop.
    """

    channel = CHANNEL_KEETA
    #: There is no TLS profile that helps: the block is the missing per-request
    #: `mtgsig`, not the ClientHello, so impersonation is beside the point.
    uses_tls_impersonation = False

    # ── sales (in-page JSON → DTOs) ─────────────────────────────────────────
    def parse_orders(self, payload: dict[str, Any]) -> list[StandardOrder]:
        """Parse an in-page `getOrders` response into orders with line items.

        The response nests the orders at `data.list[]`, each an envelope of
        `baseOrder` / `merchantOrder` / `products[]` / `feeDtl.merchantFee`;
        `_extract_rows` flattens that. One order becomes one `StandardOrder`
        with per-line `GRAIN_LINE` items, deduped on the order id.

        Dedupe is on the order id **alone**, not `(outlet, order id)`: the tree
        walk in `_extract_rows` also surfaces each envelope's nested `baseOrder`
        sub-dict as a standalone row (it carries `orderViewId`, so it satisfies
        `_looks_like_row`), and that phantom has no `shopId` — so an
        outlet-qualified key let it through as a second, empty, null-outlet copy
        of every order. The flattened envelope is always emitted before its
        phantom, and a Keeta `orderViewId` is globally unique, so keeping the
        first row per id keeps the rich one and drops the phantom.
        """
        orders: list[StandardOrder] = []
        seen: set[str] = set()
        for row in _extract_rows(payload):
            order = self._order_from(row)
            if order is None:
                continue
            if order.external_order_id in seen:
                continue
            seen.add(order.external_order_id)
            orders.append(order)
        return orders

    def _order_from(self, row: dict[str, Any]) -> StandardOrder | None:
        order_id = _first_text(row, _ORDER_ID_KEYS)
        if not order_id:
            return None
        placed_at = _first_datetime(row, _ORDER_DATE_KEYS)
        business_date = (
            placed_at.date() if placed_at else _first_date(row, _ORDER_DATE_KEYS)
        )

        gross_sales = _first_money(row, _GROSS_KEYS)
        net_sales = _first_money(row, ("netSales", "net_sales"))
        if net_sales is None:
            net_sales = gross_sales
        net_payable = _first_money(
            row,
            (
                "netPayable",
                "net_payable",
                "settleAmount",
                "settlementAmount",
                "merchantReceivable",
                "netAmount",
                "earnings",
                "total",
            ),
        )

        return StandardOrder(
            external_order_id=order_id,
            external_outlet_id=_first_text(row, _OUTLET_ID_KEYS),
            business_date=_date_str(business_date),
            placed_at=placed_at,
            status=_decode_status(
                _first_text(row, ("status", "orderStatus", "order_status"))
            ),
            currency=_first_text(row, _CURRENCY_KEYS) or "AED",
            customer_name=_customer_name(row),
            customer_phone=_customer_phone(row),
            gross_sales=gross_sales,
            net_sales=net_sales,
            commission_amount=_abs_money(
                _first_money(
                    row,
                    ("commission", "commissionAmount", "commissionFee", "brokerage"),
                )
            ),
            payment_fee=_abs_money(
                _first_money(
                    row, ("paymentFee", "transactionFee", "bankTransactionFee")
                )
            ),
            delivery_fee=_abs_money(
                _first_money(row, ("deliveryFee", "deliveryAmount"))
            ),
            vat_amount=_abs_money(_first_money(row, ("vat", "vatAmount", "taxAmount"))),
            cancellation_fee=_abs_money(
                _first_money(row, ("cancellationFee", "cancelFee"))
            ),
            net_payable=net_payable,
            statement_id=_first_text(row, _STATEMENT_ID_KEYS),
            items=self._items_from(row, order_id),
            raw=row,
        )

    def _items_from(
        self, row: dict[str, Any], order_id: str
    ) -> list[StandardOrderItem]:
        items: list[StandardOrderItem] = []
        business_date = _date_str(_first_date(row, _ORDER_DATE_KEYS))
        for index, item in enumerate(_nested_items(row), start=1):
            item_name = _first_text(
                item, ("itemName", "item_name", "name", "spuName", "skuName")
            )
            if not item_name:
                continue
            item_key = (
                _first_text(item, ("itemId", "item_id", "skuId", "spuId", "id"))
                or item_name
            )
            gross = _first_money(item, ("totalAmount", "totalPrice", "amount", "price"))
            modifier_sources = _item_modifier_sources(item)
            items.append(
                StandardOrderItem(
                    source_key=f"{order_id}:{item_key}:{index}",
                    grain=GRAIN_LINE,
                    item_name=item_name,
                    category_name=_first_text(item, ("categoryName", "category_name")),
                    quantity=_first_money(item, ("quantity", "qty", "count")),
                    unit_price=_first_money(item, ("unitPrice", "price")),
                    gross_sales=gross,
                    net_sales=gross,
                    amount_is_known=gross is not None,
                    modifiers=expand_modifiers(modifier_sources),
                    modifiers_text=_mods_text(item),
                    business_date=business_date,
                )
            )
        return items

    # ── finance (in-page JSON → DTOs, where determinable) ───────────────────
    def parse_finance(self, payload: Any) -> FinanceResult:
        """Parse an in-page finance-download response into statements + payouts.

        Keeta's finance surface hands back a list of download *tasks*, and the
        settled figures live inside the commission-invoice PDF each task points
        at — not in the JSON. When the worker has already resolved those into
        invoice-shaped rows (carrying `statementId`, `settleAmount`, period
        dates), this maps them; when the payload is only task metadata, there is
        nothing to settle yet, so it returns empty with a note rather than
        inventing zeros.

        Two payload shapes reach here, distinguished by whether they carry file
        bytes (`_has_finance_file`):

        - **File payloads** — the bootstrap worker downloaded a presigned billing
          report (`bill_xlsx_b64`) or commission invoice (`invoice_zip_b64` /
          `invoice_pdf_b64`) and pushed the bytes plus the statement identity
          (`statement_id`/`taskViewId`, period, `shopId`, `fileScene`). The whole
          payload is one statement: its XLSX becomes per-order lines, its ZIP/PDF
          is archived and stamped onto the statement.
        - **Row-tree payloads** — a bare finance-list JSON (legacy / already-
          resolved rows). `_extract_rows` walks it into statement + payout rows.
        """
        if isinstance(payload, dict) and _has_finance_file(payload):
            statement = self._statement_from(payload, 1)
            if statement is None:
                return FinanceResult(
                    statements=[],
                    payouts=[],
                    truncation_note=(
                        "Keeta finance file payload carried bytes but no statement "
                        "id (taskViewId / statement_id / time) — cannot key it."
                    ),
                )
            payouts = self._bill_payouts_from(payload, statement.statement_id)
            return FinanceResult(statements=[statement], payouts=payouts)

        rows = _extract_rows(payload)
        statements: list[StandardStatement] = []
        payouts: list[StandardPayout] = []
        seen_statements: set[str] = set()
        seen_transfers: set[str] = set()

        for index, row in enumerate(rows, start=1):
            statement = self._statement_from(row, index)
            if statement is not None and statement.statement_id not in seen_statements:
                seen_statements.add(statement.statement_id)
                statements.append(statement)
            payout = self._payout_from(row)
            if payout is not None and payout.transfer_id not in seen_transfers:
                seen_transfers.add(payout.transfer_id)
                payouts.append(payout)

        if not statements and not payouts:
            return FinanceResult(
                statements=[],
                payouts=[],
                truncation_note=(
                    "No settled Keeta finance rows in payload — the commission "
                    "invoice PDF (behind each download task's URL) carries the "
                    "figures; the in-page worker must resolve it before parse_finance."
                ),
            )
        return FinanceResult(statements=statements, payouts=payouts)

    def _statement_from(
        self, row: dict[str, Any], index: int
    ) -> StandardStatement | None:
        statement_id = _first_text(row, _STATEMENT_ID_KEYS + ("taskViewId",))
        if not statement_id:
            # Monthly commission invoices carry a `time` (YYYYMM) but no id of
            # their own — synthesise a stable one so the statement upserts.
            time_val = _get_value(row, "time")
            if time_val not in (None, "", 0):
                statement_id = f"KEETA_COMMISSION_{str(time_val).strip()}"
        if not statement_id:
            return None
        period_start = _first_date(
            row,
            ("periodStart", "period_start", "startDate", "start_date", "billStartDate"),
        )
        period_end = _first_date(
            row,
            (
                "periodEnd",
                "period_end",
                "endDate",
                "end_date",
                "billEndDate",
                "statementDate",
            ),
        )
        if period_start is None and period_end is None:
            # File payloads state the period as `displayTimeText`
            # ("15 Aug 2026 ~ 22 Aug 2026") or, for the monthly invoice, `time`.
            display_start, display_end = _period_from_display(
                _get_value(row, "displayTimeText")
            )
            if display_start or display_end:
                period_start, period_end = display_start, display_end
            else:
                period_start, period_end = _period_from_yyyymm(_get_value(row, "time"))
        currency = _first_text(row, _CURRENCY_KEYS) or "AED"

        # Archive the commission invoice (zip/pdf) when the worker sent its bytes;
        # the billing-report xlsx is data, not a VAT document, so it is parsed
        # into lines rather than archived.
        stored = None
        if _get_value(row, "invoice_zip_b64") or _get_value(row, "invoice_pdf_b64"):
            stored = _archive_keeta_invoice(row, statement_id)

        return StandardStatement(
            statement_id=statement_id,
            period_start=_date_str(period_start),
            period_end=_date_str(period_end),
            payment_due_date=_date_str(
                _first_date(
                    row,
                    ("paymentDueDate", "payment_due_date", "payDate", "paymentDate"),
                )
            ),
            gross_sales=_first_money(
                row,
                (
                    "grossSales",
                    "gross_sales",
                    "totalSales",
                    "orderAmount",
                    "totalAmount",
                ),
            ),
            net_payable=_first_money(
                row,
                (
                    "netPayable",
                    "net_payable",
                    "settleAmount",
                    "settlementAmount",
                    "paymentAmount",
                ),
            ),
            total_fees=_abs_money(
                _first_money(
                    row, ("totalFees", "total_fees", "feeAmount", "commissionAmount")
                )
            ),
            total_vat=_abs_money(_first_money(row, ("vat", "vatAmount", "taxAmount"))),
            currency=currency,
            external_outlet_id=_first_text(row, _OUTLET_ID_KEYS),
            invoice_object_key=stored.object_key if stored else None,
            invoice_content_type=stored.content_type if stored else None,
            invoice_original_filename=stored.original_filename if stored else None,
            invoice_fetched_at=stored.fetched_at if stored else None,
            invoice_attachments=stored.attachments if stored else None,
            lines=self._statement_lines_from(row, statement_id, index),
            raw=_strip_blobs(row),
        )

    def _statement_lines_from(
        self, row: dict[str, Any], statement_id: str, index: int
    ) -> list[StandardStatementLine]:
        # A billing-report xlsx is the per-order detail: decode it and walk the
        # "Order Summary" sheet into gross/commission/bank-fee/net lines.
        xlsx_b64 = _get_value(row, "bill_xlsx_b64")
        if xlsx_b64:
            try:
                xlsx_bytes = base64.b64decode(xlsx_b64)
            except Exception:  # noqa: BLE001 — a bad blob yields no lines, not a crash
                logger.warning(
                    "keeta: invalid bill_xlsx_b64 for %s — no lines", statement_id
                )
                return []
            return _parse_bill_xlsx(xlsx_bytes, statement_id)

        amount = _first_money(
            row,
            (
                "amount",
                "feeAmount",
                "commissionAmount",
                "adjustmentAmount",
                "settleAmount",
            ),
        )
        if amount is None:
            return []
        source_key = (
            _first_text(row, ("lineId", "line_id", "id")) or f"{statement_id}:{index}"
        )
        return [
            StandardStatementLine(
                source_key=source_key,
                statement_id=statement_id,
                transfer_id=_first_text(row, _TRANSFER_ID_KEYS),
                external_order_id=_first_text(row, _ORDER_ID_KEYS),
                line_date=_date_str(
                    _first_date(
                        row, ("lineDate", "date", "businessDate", "statementDate")
                    )
                ),
                line_type=_normalize_status(
                    _first_text(row, ("lineType", "type", "category"))
                )
                or "statement_line",
                fee_category=_normalize_status(
                    _first_text(row, ("feeCategory", "fee_category", "category"))
                ),
                description=_first_text(row, ("description", "desc", "remark", "name")),
                amount=amount,
                currency=_first_text(row, _CURRENCY_KEYS) or "AED",
            )
        ]

    def _bill_payouts_from(
        self, row: dict[str, Any], statement_id: str
    ) -> list[StandardPayout]:
        """Weekly settlement payouts from a bill xlsx's "Invoice Details" sheet.

        Only a bill-report payload (carrying `bill_xlsx_b64`) has a settlement
        total to parse; a commission-invoice zip/pdf payload has none, so this
        returns `[]`. The payouts are keyed and coupled to `statement_id` (the
        same id the weekly statement uses) inside `_parse_bill_payouts`.
        """
        xlsx_b64 = _get_value(row, "bill_xlsx_b64")
        if not xlsx_b64:
            return []
        try:
            xlsx_bytes = base64.b64decode(xlsx_b64)
        except Exception:  # noqa: BLE001 — a bad blob yields no payouts, not a crash
            logger.warning(
                "keeta: invalid bill_xlsx_b64 for %s — no payouts", statement_id
            )
            return []
        return _parse_bill_payouts(
            xlsx_bytes,
            statement_id,
            _first_text(row, _OUTLET_ID_KEYS),
            _first_text(row, ("taskViewId",)),
        )

    def _payout_from(self, row: dict[str, Any]) -> StandardPayout | None:
        transfer_id = _first_text(row, _TRANSFER_ID_KEYS)
        if not transfer_id:
            return None
        return StandardPayout(
            transfer_id=transfer_id,
            statement_id=_first_text(row, _STATEMENT_ID_KEYS),
            transfer_date=_date_str(
                _first_date(
                    row,
                    (
                        "transferDate",
                        "transfer_date",
                        "paymentDate",
                        "payDate",
                        "settleDate",
                    ),
                )
            ),
            payment_due_date=_date_str(
                _first_date(row, ("paymentDueDate", "payment_due_date"))
            ),
            transfer_amount=_first_money(
                row,
                (
                    "transferAmount",
                    "paymentAmount",
                    "payoutAmount",
                    "settleAmount",
                    "amount",
                ),
            ),
            transfer_status=_normalize_status(
                _first_text(row, ("status", "paymentStatus", "transferStatus"))
            ),
            payment_reference=_first_text(
                row, ("reference", "paymentReference", "bankReference")
            ),
            currency=_first_text(row, _CURRENCY_KEYS) or "AED",
        )

    # ── the httpx interface — deliberately inert for Keeta ───────────────────
    async def fetch_sales(
        self, session: LoadedSession, *, since: datetime, until: datetime
    ) -> SalesResult:
        """Not reachable over httpx — see the module docstring. Raises so a
        misrouted Keeta session on the ingest loop fails loudly."""
        raise AggregatorUnavailableError(_UNAVAILABLE)

    async def fetch_statements(
        self, session: LoadedSession, *, since: datetime, until: datetime
    ) -> StatementsResult:
        raise AggregatorUnavailableError(_UNAVAILABLE)

    async def fetch_payouts(
        self, session: LoadedSession, *, since: datetime, until: datetime
    ) -> PayoutsResult:
        raise AggregatorUnavailableError(_UNAVAILABLE)


#: The module-level singleton, matching the careem/foodics/grubops providers —
#: it is stateless (the payload is passed in per call), so sharing it is free.
provider = KeetaClient()
