"""Talabat, as its partner portal's two GraphQLs answer.

The hardest of the five, on two counts. First, it is two APIs, not one: orders
live behind a supergraph, finance behind a separate gateway, and the two speak
different request envelopes.

- Sales: `POST https://vm-supergraph.eu.prd.portal.restaurant/graphql`. There is
  no "give me the orders" query — Talabat only exports orders as an async CSV
  job. So a pull is three GraphQL calls (`EstimateExportSize` → `RequestExport`
  → poll `ListExports` until the export completes and carries a `downloadUrl`),
  then an authenticated GET of that CSV, which is parsed into orders and their
  line items. This mirrors the browser's Report Builder flow exactly; the only
  thing the browser did that this cannot is *discover* the store ids by scraping
  the page, so those are read from the session (see `_order_account_ids`).
- Finance: `POST https://vagw-api.eu.prd.portal.restaurant/query`. Payouts
  (`ListPayouts`) and settlement metadata (`ListAdditionalStatements`) are real
  paginated queries, so finance is a straight map. The `accounts` descriptors
  these queries key on were captured off the network by the bootstrap and are
  read back from the session (see `_finance_accounts`).

Second, the bot wall. Talabat fronts everything with PerimeterX, which
fingerprints the TLS ClientHello, so `uses_tls_impersonation` is set and, where
`curl_cffi` is installed, the base sends a Chrome ClientHello. PerimeterX also
answers a flagged request not only with a 403 but sometimes a 200 whose body is
a "press & hold" / captcha challenge, so `_is_auth_failure` is overridden to
read the body as well as the status — a challenge is a dead session, not data.

Auth is a bearer `accessToken`. In the browser it is read from the `accessToken`
cookie first and localStorage second; here it is read from the session's cookies
first and its tokens second (`_access_token`), and sent as `Authorization:
Bearer …` on every call plus an `x-global-entity-id: tb_ae` header. Entity
defaults to `TB_AE` (Talabat UAE) but is read off the session
(`tokens["global_entity_id"]`, injected from `aggregator_account.extras` by
`session_store.enrich_session`) so a non-UAE entity needs no code change.

Money is taken verbatim and left null where a column is absent (None is
"unknown", not zero — a blank commission cell is not a free order), every order
keeps its `raw`, and a window the portal could not export in full sets a
`truncation_note` rather than implying it covered everything.
"""

from __future__ import annotations

import asyncio
import csv
import io
import logging
import re
import zipfile
from dataclasses import replace
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.aggregator import CHANNEL_TALABAT, GRAIN_LINE
from app.services.aggregators.normalized import (
    PayoutsResult,
    SalesResult,
    StandardModifier,
    StandardOrder,
    StandardOrderItem,
    StandardPayout,
    StandardStatement,
    StandardStatementLine,
    StandardStatusEvent,
    StatementsResult,
)
from app.services.aggregators.session_store import LoadedSession
from app.services.providers._agg_parse import first_present as _first
from app.services.providers._agg_parse import parse_money as _money
from app.services.providers.aggregator_base import (
    AggregatorAuthError,
    AggregatorUnavailableError,
    BaseAggregatorClient,
)

logger = logging.getLogger(__name__)

# ── Endpoints and entity ──────────────────────────────────────────────────────
_ORDERS_GRAPHQL = "https://vm-supergraph.eu.prd.portal.restaurant/graphql"
_FINANCE_GRAPHQL = "https://vagw-api.eu.prd.portal.restaurant/query"
#: The default Talabat global entity (UAE). The live value is read off the
#: session (`tokens["global_entity_id"]`, populated from `aggregator_account`
#: `.extras` by `session_store.enrich_session`); this is only the fallback when
#: the account carries none, so behaviour is unchanged until an operator sets it.
_GLOBAL_ENTITY_ID = "TB_AE"

# ── Report Builder export constants (orders) ──────────────────────────────────
_EXPORT_TYPE_ORDERS = "EXPORT_TYPE_ORDERS_LIST"
_EXPORT_ACCOUNT_TYPE_VENDOR = "EXPORT_ACCOUNT_TYPE_VENDOR"
_DELIVERY_METHOD_DIRECT = "DELIVERY_METHOD_DIRECT"
_EXPORT_FORMAT_CSV = "CSV"
_EXPORT_STATUS_COMPLETED = "EXPORT_STATUS_COMPLETED"
_EXPORT_STATUS_TERMINAL_BAD = {
    "EXPORT_STATUS_FAILED",
    "EXPORT_STATUS_CANCELLED",
    "EXPORT_STATUS_EXPIRED",
}
_EXPORT_POLL_TIMEOUT_SECONDS = 180
_EXPORT_POLL_INTERVAL_SECONDS = 5
#: Year-long order exports can be several MB; the default 20s aggregator timeout
#: is too tight once PerimeterX has let the download start.
_CSV_DOWNLOAD_TIMEOUT_SECONDS = 120.0

# ── Finance pagination ────────────────────────────────────────────────────────
_FINANCE_PAGE_SIZE = 100
#: Hard ceiling so a stable `nextPageToken` (an API bug, or a session that has
#: quietly expired into a redirect) cannot spin the loop forever against a live
#: portal — ported verbatim from the Playwright scraper's guard.
_MAX_FINANCE_PAGES = 500

# ── GraphQL documents (ported verbatim from the Playwright exporter) ───────────
_ESTIMATE_EXPORT_SIZE_QUERY = """
query EstimateExportSize($input: EstimateExportSizeReq!) {
  estimateExportSize(input: $input) {
    withinLimits
    estimatedFileCount
    estimatedRowCount
    __typename
  }
}
""".strip()

_REQUEST_EXPORT_MUTATION = """
mutation RequestExport($input: RequestExportReq!) {
  requestExport(input: $input) {
    exportId
    status
    __typename
  }
}
""".strip()

_LIST_EXPORTS_QUERY = """
query ListExports($input: ListExportsReq!) {
  listExports(input: $input) {
    exports {
      exportId
      name
      exportType
      status
      requestedAt
      startedAt
      expiresAt
      downloadUrl
      columns {
        paths
        __typename
      }
      filters {
        from
        to
        accountIds
        accountType
        __typename
      }
      __typename
    }
    nextPageToken
    prevPageToken
    __typename
  }
}
""".strip()

_LIST_PAYOUTS_QUERY = """
query ListPayouts($params: ListPayoutsRequest!) {
  finances {
    listPayouts(input: $params) {
      nextPageToken
      prevPageToken
      payouts {
        payoutId: id
        payoutAmount: netPayout
        payoutCurrency: currency
        payoutOrders: ordersCount
        at: paymentDateLocal
        status: payoutStatus
        payoutAttachments: attachments
        payoutAccount: account {
          grid
          billingParentId
          chainId
          __typename
        }
        invoices {
          invoiceId: id
          invoiceAmount: totalPayout
          invoiceCurrency: currency
          invoiceOrders: ordersCount
          processedDate
          invoiceAttachments: attachments
          period: earningsPeriod {
            from: invoiceStartDate
            to: invoiceEndDate
            __typename
          }
          invoiceAccount: account {
            grid
            billingParentId
            chainId
            __typename
          }
          __typename
        }
        __typename
      }
      __typename
    }
    __typename
  }
}
""".strip()

_LIST_ADDITIONAL_STATEMENTS_QUERY = """
query ListAdditionalStatements($params: ListAdditionalStatementsRequest!) {
  finances {
    listAdditionalStatements(input: $params) {
      nextPageToken
      prevPageToken
      additionalStatements {
        statementId
        statementType
        amountGross
        currency
        statementDate
        globalEntityId
        attachments {
          path
          type
          __typename
        }
        account {
          grid
          billingParentId
          chainId
          __typename
        }
        __typename
      }
      __typename
    }
    __typename
  }
}
""".strip()

_BULK_STATEMENT_COUNTS_QUERY = """
query GetBulkAdditionalStatementDownloadCounts($params: GetBulkAdditionalStatementDownloadCountsRequest!) {
  finances {
    getBulkAdditionalStatementDownloadCounts(input: $params) {
      fileCounts {
        totalFilesCount
        fileCountsDetails {
          filesCount
          fileFormat
          __typename
        }
        __typename
      }
      fileLimits {
        directDownloadLimit
        emailDownloadLimit
        __typename
      }
      __typename
    }
    __typename
  }
}
""".strip()

_BULK_STATEMENT_DOWNLOAD_QUERY = """
query RequestBulkDownloadAdditionalStatements($params: BulkDownloadAdditionalStatementsRequest!) {
  finances {
    bulkDownloadAdditionalStatements(input: $params) {
      downloadUrl
      __typename
    }
    __typename
  }
}
""".strip()


# ── value helpers ─────────────────────────────────────────────────────────────


#: Markers that mean a Talabat GraphQL `errors` array is really a dead session,
#: not a transient upstream hiccup. The gateway answers 200 and buries the auth
#: failure of a downstream service (`vp-report-builder`) in the errors — an expired
#: bearer surfaces as a 401 SUBREQUEST_HTTP_ERROR or a TOKEN_EXPIRED malformed
#: response — so the HTTP status the base inspects never sees it. Matched against
#: the serialised errors, lower-cased; kept auth-specific (no bare "401", which a
#: store id could carry) so a real VALIDATION_ERROR still reads as unavailable and
#: retries rather than forcing a re-login.
_GRAPHQL_AUTH_MARKERS = (
    "token_expired",
    "token expired",
    "unauthorized",
    "unauthenticated",
)


def _graphql_errors_are_auth(errors: Any) -> bool:
    """Whether a GraphQL `errors` payload signals an expired/invalid session."""
    blob = str(errors).lower()
    return any(marker in blob for marker in _GRAPHQL_AUTH_MARKERS)


def _parse_dt(value: Any) -> datetime | None:
    """The CSV's `Order received at` (`YYYY-MM-DD HH:MM[:SS]`) as a datetime."""
    if not value or not isinstance(value, str):
        return None
    for pattern in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value.strip(), pattern)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
    except ValueError:
        return None


#: The export's status-timeline columns, in the order an order moves through
#: them, paired with the normalized status word each maps to. `_orders_from_csv`
#: walks this to build `StandardOrder.status_events`: a cell with a value emits
#: one event (dated by `_parse_dt`), sequenced by its position here. An empty or
#: absent cell is a step that has not happened, so it emits nothing.
_TIMELINE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Order received at", "received"),
    ("Accepted at", "accepted"),
    ("Ready to pick up at", "ready"),
    ("Rider near pickup at", "rider_near_pickup"),
    ("In delivery at", "in_delivery"),
    ("Delivered at", "delivered"),
    ("Cancelled at", "cancelled"),
)


def _status_events_from_row(row: dict[str, Any]) -> list[StandardStatusEvent]:
    """The order's status trace, built from the export's timeline columns.

    One `StandardStatusEvent` per timeline column that carries a value, in the
    fixed lifecycle order of `_TIMELINE_COLUMNS` (so `sequence` reflects that
    order, not the sparse subset actually present). The cell is parsed to a
    naive Dubai datetime; a column that is empty or absent is skipped, so a
    still-open order carries only the steps it has reached.
    """
    events: list[StandardStatusEvent] = []
    for sequence, (column, status) in enumerate(_TIMELINE_COLUMNS):
        cell = row.get(column)
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            continue
        events.append(
            StandardStatusEvent(
                status=status,
                at=_parse_dt(cell),
                sequence=sequence,
            )
        )
    return events


def _parse_date(value: Any) -> date | None:
    """A finance date (`Aug 26, 2026` / `2026-08-26` / `26/08/2026`) as a date."""
    if not value or not isinstance(value, str):
        return None
    for pattern in ("%b %d, %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value.strip(), pattern).date()
        except ValueError:
            continue
    return None


def _date_str(value: Any) -> str | None:
    """A finance date normalised to `YYYY-MM-DD`, or None."""
    parsed = _parse_date(value)
    return parsed.isoformat() if parsed else None


def _split_balanced(value: str) -> list[str]:
    """Comma-split a string, keeping commas inside (…) as part of the same token.

    `"1 Burger (No pickle, Extra cheese), 2 Fries"` → two tokens, not three.
    """
    tokens: list[str] = []
    current: list[str] = []
    depth = 0
    for ch in value:
        if ch == "(":
            depth += 1
            current.append(ch)
        elif ch == ")":
            depth = max(0, depth - 1)
            current.append(ch)
        elif ch == "," and depth == 0:
            tok = "".join(current).strip()
            if tok:
                tokens.append(tok)
            current = []
        else:
            current.append(ch)
    last = "".join(current).strip()
    if last:
        tokens.append(last)
    return tokens


def _extract_item_modifiers(name: str) -> tuple[str, list[str]]:
    """Strip trailing parenthetical or `+ addon` chains from an item name token.

    Returns `(cleaned_name, list_of_modifier_texts)`. Handles:
    - Parenthetical: `"Burger (No pickle, Extra cheese)"` →
      `("Burger", ["No pickle", "Extra cheese"])`
    - Plus-addon chain: `"Burger + Extra sauce + No pickle"` →
      `("Burger", ["Extra sauce", "No pickle"])`
    - Both combined: parenthetical is extracted first, then any remaining `+` chains.
    """
    mods: list[str] = []
    # Parenthetical group at end
    paren = re.search(r"\(([^)]+)\)\s*$", name)
    if paren:
        inner = [s.strip() for s in paren.group(1).split(",") if s.strip()]
        mods.extend(inner)
        name = name[: paren.start()].strip()
    # Plus-addon chain
    if " + " in name:
        parts = name.split(" + ")
        name = parts[0].strip()
        mods.extend(p.strip() for p in parts[1:] if p.strip())
    return name, mods


def _parse_items_text(value: str | None) -> list[tuple[Decimal, str]]:
    """The CSV's free-text `Order Items` field split into (qty, name) lines.

    A cell is a delimiter-joined list of `"<qty> <name>"` tokens; a token that
    does not start with a number is taken as quantity 1. Ported from the
    exporter's `_parse_talabat_order_items_text`.

    Newlines and semicolons are always safe item separators. Commas are split
    only when outside balanced parentheses, so modifier lists like
    `"1 Burger (No pickle, Extra cheese), 2 Fries"` produce two items rather
    than three broken fragments.
    """
    if not value:
        return []
    # Newlines / semicolons are unambiguous item separators.
    if "\n" in value or ";" in value:
        raw_tokens = [t.strip() for t in re.split(r"[\n;]+", value) if t.strip()]
    else:
        raw_tokens = _split_balanced(value)
    parsed: list[tuple[Decimal, str]] = []
    for token in raw_tokens:
        match = re.match(r"^(?P<qty>\d+(?:\.\d+)?)\s+(?P<name>.+)$", token)
        if match:
            parsed.append((Decimal(match.group("qty")), match.group("name").strip()))
        else:
            parsed.append((Decimal(1), token))
    return parsed


# ── statement bundle helpers (ported from automation parsers.py) ──────────────

_STATEMENT_PERIOD_PATTERN = re.compile(
    r"(?P<brand>.+?)\s*-\s*(?P<start>\d{2}/\d{2}/\d{4})"
    r"\s*-\s*(?P<end>\d{2}/\d{2}/\d{4})$"
)


def _bundle_period(title: object) -> tuple[date | None, date | None]:
    """Parse the sheet title `"Brand - DD/MM/YYYY - DD/MM/YYYY"` into dates."""
    if not isinstance(title, str):
        return None, None
    m = _STATEMENT_PERIOD_PATTERN.match(title.strip())
    if not m:
        return None, None
    return _parse_date(m.group("start")), _parse_date(m.group("end"))


def _bundle_headers(row: tuple[object, ...]) -> dict[str, int]:
    """Normalised column-name → index map for an xlsx header row."""
    return {
        str(v).strip(): i
        for i, v in enumerate(row)
        if isinstance(v, str) and str(v).strip()
    }


def _bundle_cell(
    row: tuple[object, ...], headers: dict[str, int], label: str
) -> object | None:
    idx = headers.get(label)
    return row[idx] if idx is not None and idx < len(row) else None


def _bundle_is_order_id(value: object) -> bool:
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    return str(value or "").strip().isdigit()


def _bundle_order_id_str(value: object) -> str:
    if isinstance(value, float):
        return str(int(value))
    return str(value or "").strip()


def _bundle_money(
    row: tuple[object, ...], headers: dict[str, int], col: str
) -> Decimal:
    """A Talabat xlsx cell as Decimal; zero for absent / blank / bad values."""
    return _money(_bundle_cell(row, headers, col)) or Decimal("0")


def _month_windows(from_date: date, to_date: date) -> list[tuple[date, date]]:
    """Monthly sub-windows covering `[from_date, to_date]` inclusive.

    Used when the file count exceeds the portal's direct-download limit and the
    request must be chunked by calendar month — ported from automation exports.py.
    """
    windows: list[tuple[date, date]] = []
    current = date(from_date.year, from_date.month, 1)
    while current <= to_date:
        if current.month == 12:
            next_month = date(current.year + 1, 1, 1)
        else:
            next_month = date(current.year, current.month + 1, 1)
        windows.append(
            (max(from_date, current), min(to_date, next_month - timedelta(days=1)))
        )
        current = next_month
    return windows


#: The DeliveryHero vendor-api that backs Talabat's menu-management console (a
#: different host from the order/finance GraphQL). UAE platform. Verified live from
#: the VM session 2026-09-01 — the stored session carries the `authorization`
#: bearer + `x-global-entity-id` the SPA sends, and request_json's TLS impersonation
#: passes PerimeterX, so the menu reads with the same session the sales ingest uses.
_MENU_API = "https://vendor-api-gdp-ae.me.restaurant-partners.com/api/5/platforms/TB_AE"
#: Origin of the vendor-api host — statement-attachment paths are often relative
#: (`/api/5/platforms/.../file`) and must be fetched here, not against GraphQL.
_VENDOR_API_ORIGIN = "https://vendor-api-gdp-ae.me.restaurant-partners.com"
#: DeliveryHero Vendor Time Service — opening hours (not on the menu API).
_VTS = "https://vts.eu.restaurant-partners.com/opening-times/v1"


class TalabatClient(BaseAggregatorClient):
    channel = CHANNEL_TALABAT
    uses_tls_impersonation = True
    impersonate_target = "chrome"

    # ── PerimeterX-aware auth-failure detection ───────────────────────────────
    def _is_auth_failure(self, response: Any) -> bool:
        """A dead session — a 401/403, or a PerimeterX bot challenge.

        PerimeterX flags a bad TLS/cookie fingerprint with a 403, but also
        sometimes with a 200 whose body is the "press & hold" / captcha
        interstitial rather than the expected JSON. Either way the session can
        only be revived by a browser bootstrap, so both map to an auth failure
        and never to a retry.
        """
        if getattr(response, "status_code", None) in (401, 403):
            return True
        try:
            body = (getattr(response, "text", "") or "")[:4000].lower()
        except Exception:  # noqa: BLE001 - an unreadable body is not a challenge
            return False
        markers = (
            "px-captcha",
            "perimeterx",
            "_pxhd",
            "press & hold",
            "press and hold",
            "access to this page has been denied",
            "please enable js",
            "human verification",
        )
        return any(marker in body for marker in markers)

    def build_headers(
        self, session: LoadedSession, extra: dict[str, str] | None = None
    ) -> dict[str, str]:
        """Base headers, but with a FRESH `Authorization` bearer.

        The captured `header_profile` carries an `Authorization: Bearer <token>` from
        login time, and the base sends it verbatim. A heal that refreshes the
        `accessToken` cookie but not the profile then leaves the REST menu read
        replaying a stale bearer — the vendor-api answers `401 Authentication failed`
        with the session still marked `live` (found 2026-09-03: a fresh-cookie token
        as the bearer returns 200 on the same call). So overwrite any profiled
        Authorization with the current cookie token; the GraphQL path already did this
        per-call, this makes the REST path do it too."""
        headers = super().build_headers(session, extra)
        token = self._access_token(session)
        if token:
            for key in [k for k in headers if k.lower() == "authorization"]:
                del headers[key]
            headers["Authorization"] = f"Bearer {token}"
        return headers

    # ── session-sourced credentials ───────────────────────────────────────────
    @staticmethod
    def _global_entity_id(session: LoadedSession) -> str:
        """The Talabat global entity id for this session.

        Read from `tokens["global_entity_id"]` (injected from
        `aggregator_account.extras` by `session_store.enrich_session`), falling
        back to `_GLOBAL_ENTITY_ID` (`TB_AE`) when the account carries none — so
        the request is byte-identical to before until an operator sets it.
        """
        value = (session.tokens or {}).get("global_entity_id")
        if isinstance(value, str) and value.strip():
            return value.strip()
        return _GLOBAL_ENTITY_ID

    @staticmethod
    def _access_token(session: LoadedSession) -> str | None:
        """The bearer `accessToken`, from the session's cookies then tokens.

        The browser read it from the `accessToken` cookie first and the
        `authentication` localStorage blob second; the bootstrap captures both
        into the session, so this reads `cookies["accessToken"]` first, then the
        common token keys.
        """
        cookie_token = (session.cookies or {}).get("accessToken")
        if isinstance(cookie_token, str) and cookie_token.strip():
            return cookie_token.strip()
        tokens = session.tokens or {}
        for key in ("accessToken", "access_token", "bearer", "token"):
            value = tokens.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        return None

    @staticmethod
    def _finance_accounts(session: LoadedSession) -> list[dict[str, Any]]:
        """The `accounts` descriptors the finance queries key on.

        The browser captured these off the finance page's own network traffic;
        the bootstrap stores them under one of these token keys. Each is a
        `{grid, billingParentId, chainId}` dict the gateway echoes back.
        """
        tokens = session.tokens or {}
        for key in ("finance_accounts", "accounts", "financeAccounts"):
            value = tokens.get(key)
            if isinstance(value, list) and value:
                return [a for a in value if isinstance(a, dict)]
        return []

    def _order_account_ids(self, session: LoadedSession) -> list[str]:
        """The store ids the order export scopes to.

        The browser scraped these from the Report Builder store picker; here
        they come from the session's tokens (`account_ids`/`store_ids`), falling
        back to the `grid`s of the finance accounts, which are the same numeric
        store ids.
        """
        tokens = session.tokens or {}
        for key in ("account_ids", "store_ids", "accountIds"):
            value = tokens.get(key)
            if isinstance(value, list) and value:
                return [str(v) for v in value if v is not None]
        grids = [
            str(a["grid"])
            for a in self._finance_accounts(session)
            if a.get("grid") is not None
        ]
        return grids

    @staticmethod
    def _account_ids_in(tokens: dict) -> bool:
        for key in ("account_ids", "store_ids", "accountIds"):
            value = tokens.get(key)
            if isinstance(value, list) and value:
                return True
        return False

    # ── catalog (catalog sync) ───────────────────────────────────────────────
    # Verified live from the VM session 2026-09-01:
    #   {_MENU_API}/vendors/{v}/catalogs
    #     -> {catalogs:[{id, name, categories:[{id, name}]}], ...}
    #   {_MENU_API}/vendors/{v}/catalogs/{catalogId}/categories/{categoryId}/products
    #     -> [{id, name, description, unitPrice, availability:{available}, active,
    #          productOptionIds, isVariation, ...}]
    #
    # Create: the nested per-category products URL is GET-only (405 on POST).
    # The Add Product drawer (menuManagementV2 1.15.11, captured 2026-09-04 on
    # Karama vendor 793319) POSTs `{_MENU_API}/vendors/{v}/catalogs/products`
    # with `{name, description, unitPrice, catalogIds, category, type:"Simple",
    # active}`. Response is `{commandId}` (async command, not an immediate
    # product id). Still behind `CATALOG_SYNC_ENABLED`, dry-run default.

    async def list_catalogs(self, session: LoadedSession, vendor: str) -> Any:
        """The vendor's catalogs, each with its `categories` inline."""
        return await self.request_json(
            session, "GET", f"{_MENU_API}/vendors/{vendor}/catalogs"
        )

    async def list_category_products(
        self,
        session: LoadedSession,
        vendor: str,
        catalog_id: str,
        category_id: str,
    ) -> Any:
        """One category's products (a bare array)."""
        return await self.request_json(
            session,
            "GET",
            f"{_MENU_API}/vendors/{vendor}/catalogs/{catalog_id}"
            f"/categories/{category_id}/products",
            params={"locale": "en-AE", "sizeSupport": "true"},
        )

    async def list_product_options(self, session: LoadedSession, vendor: str) -> Any:
        """The vendor's product-option (modifier) groups — `GET /vendors/{v}/
        catalogs/product-options` (verified live 2026-09-05). A bare array; each
        group is `{id, name, names:[{locale,value}], quantity:{minimum,maximum},
        options:[{id, name, names:[...], unitPrice, active, availability}]}`. This
        is the mix-box "Options (Max N)" flavour groups a product references by
        `productOptionIds` — distinct from a SIZED_PRODUCT's `nestedProducts` sizes.
        A single group can also be fetched at `/catalogs/product-options/{id}`."""
        return await self.request_json(
            session,
            "GET",
            f"{_MENU_API}/vendors/{vendor}/catalogs/product-options",
            params={"locale": "en"},
        )

    async def get_product_detail(
        self, session: LoadedSession, vendor: str, product_id: str
    ) -> Any:
        """One product's full record, including its sizes. The category-products
        list carries only `productOptionIds`; the sizes (name + price) live in this
        detail call's `nestedProducts` (each `type:"SIZE"`). Endpoint captured live
        from the portal 2026-09-05 (`/catalogs/products/{id}?mode=READ_ONLY`).

        Localisation note (verified live 2026-09-05): Talabat's Karama menu is
        already fully bilingual — Arabic name/description AND Arabic size names are
        present and match MM. They are NOT in the READ_ONLY `names`/`descriptions`
        arrays (those come back empty); they live in the `?mode=EDITING` payload,
        each as `[{"locale":"ar-AE","value":...},{"locale":"en-AE","value":...}]`.
        `?mode=EDITING` also exposes `active` and `availability`
        ({"available":bool,"status":"COMMITTED"}) which READ_ONLY carries too.

        Writes (verified 2026-09-05, see `update_product`): **PATCH**
        `/catalogs/products/{id}` (PUT 405s). Known fields are [unitPrice,
        nestedProductOptions, names, vat, name, environmentalTax, pvcTax, posId,
        allergens, productOptionIds, netAmount, descriptions, tags, type, id,
        description, imageUrls, nutritionFacts, packagingCharge, nestedProducts,
        categories, catalogIds] — so en+ar title/desc, price and image all go here;
        there is NO `active` field on PATCH. On-shelf activation is a SEPARATE async
        endpoint (`set_availability`). Mix-box flavour choices are a
        `productOptionIds` group (Options (Max N)), not a SIZE."""
        return await self.request_json(
            session,
            "GET",
            f"{_MENU_API}/vendors/{vendor}/catalogs/products/{product_id}",
            params={"locale": "en", "mode": "READ_ONLY"},
        )

    async def create_menu_item(
        self,
        session: LoadedSession,
        vendor: str,
        *,
        name: str,
        catalog_id: str,
        category_id: str,
        price: Any,
        description: str = "",
        active: bool = False,
    ) -> Any:
        """Create one item via the partner Add Product drawer POST.

        Nested `.../catalogs/{id}/categories/{id}/products` is GET-only (405).
        Off-shelf (`active=false`) by default so a sync never goes live before
        review. Response is `{commandId}` — map the product on the next menu read.
        """
        return await self.request_json(
            session,
            "POST",
            f"{_MENU_API}/vendors/{vendor}/catalogs/products",
            json_body={
                "name": name,
                "description": description or "",
                "unitPrice": float(price),
                "catalogIds": [str(catalog_id)],
                "category": str(category_id),
                "type": "Simple",
                "active": bool(active),
            },
        )

    @staticmethod
    def _locales(en: str | None, ar: str | None) -> list[dict[str, str]]:
        """Talabat's localised-field array shape ([{locale, value}]) — the EDITING
        payload uses `ar-AE`/`en-AE`. Only non-None values are emitted."""
        out: list[dict[str, str]] = []
        if ar is not None:
            out.append({"locale": "ar-AE", "value": ar})
        if en is not None:
            out.append({"locale": "en-AE", "value": en})
        return out

    async def update_product(
        self,
        session: LoadedSession,
        vendor: str,
        product_id: str,
        *,
        name: str | None = None,
        name_ar: str | None = None,
        description: str | None = None,
        description_ar: str | None = None,
        price: Any | None = None,
        image_url: str | None = None,
    ) -> Any:
        """Edit a product — PATCH `/catalogs/products/{id}` (verified live
        2026-09-05). Only the passed fields are sent. Localised title/description
        go as BOTH the scalar (`name`/`description`) and the `names`/`descriptions`
        [{locale,value}] arrays (en-AE + ar-AE), which is how the console writes
        bilingual copy. Price is `unitPrice`; image is `imageUrls` (a list). PUT
        405s here; this is NOT how on-shelf status changes — use `set_availability`.
        """
        payload: dict[str, Any] = {}
        if name is not None:
            payload["name"] = name
            payload["names"] = self._locales(name, name_ar)
        elif name_ar is not None:
            payload["names"] = self._locales(None, name_ar)
        if description is not None:
            payload["description"] = description
            payload["descriptions"] = self._locales(description, description_ar)
        elif description_ar is not None:
            payload["descriptions"] = self._locales(None, description_ar)
        if price is not None:
            payload["unitPrice"] = float(price)
        if image_url is not None:
            payload["imageUrls"] = [image_url]
        return await self.request_raw(
            session,
            "PATCH",
            f"{_MENU_API}/vendors/{vendor}/catalogs/products/{product_id}",
            json_body=payload,
        )

    async def set_availability(
        self, session: LoadedSession, vendor: str, product_id: str, *, available: bool
    ) -> Any:
        """Bring a product on/off shelf — PUT `/catalogs/products/{id}/availability`
        `{"available": bool, "availableTimeChanges": []}` (verified live
        2026-09-05). This is the on-shelf `active` toggle, distinct from a PATCH
        edit. The call is **async**: it returns 202 and the product's `active`
        flag flips a few seconds later (empty `availableTimeChanges` = always, no
        schedule). Sending a bare `{"available": true}` is accepted (202) but does
        NOT take effect — the `availableTimeChanges` key is required."""
        return await self.request_raw(
            session,
            "PUT",
            f"{_MENU_API}/vendors/{vendor}/catalogs/products/{product_id}/availability",
            json_body={"available": bool(available), "availableTimeChanges": []},
        )

    async def get_delivery_calendars(self, session: LoadedSession, vendor: str) -> Any:
        """The vendor's DELIVERY opening-hours calendars, read server-side.

        Opening hours live on the DeliveryHero Vendor Time Service
        (`vts.eu.restaurant-partners.com`), NOT the menu API — the portal's
        Opening-Times page fetches `.../opening-times/v1/vendor/TB_AE;{v}/
        calendars/DELIVERY`. It replays fine over the same TLS-impersonating
        session + vendor-portal bearer the menu/sales reads use (verified live
        2026-09-02: 200 for all three MM vendors). Returns
        `{calendars:[{name, schedule:{openingTimesByDay:[{day, openingTimes:
        [{from, to}]}]}}]}` — `from`/`to` are minutes-from-midnight, `day` is
        0=Monday..6=Sunday (`entity/configuration.firstDOW=0`)."""
        return await self.request_json(
            session,
            "GET",
            f"{_VTS}/vendor/TB_AE;{vendor}/calendars/DELIVERY",
        )

    async def put_delivery_calendars(
        self, session: LoadedSession, vendor: str, calendar: Any
    ) -> Any:
        """Write ONE of the vendor's calendars (its whole object, from the GET).

        Verified live 2026-09-05: the write is `PUT .../calendars` (the collection,
        NOT `/calendars/DELIVERY` — that path is GET-only and 405s every write
        method), and the body is a **single calendar object** (the full shape the
        GET returns: `{id, name, schedule:{type, openingTimesByDay}, comment,
        visible, active}`), NOT an array or a `{calendars:[…]}` wrapper (both 400
        "Failed to read request"). Sending only the "Normal" calendar edits that
        one and leaves the alternative (ramadan/eid) calendars untouched. Returns
        200. Live writes are behind `CATALOG_SYNC_ENABLED` and default dry-run.
        """
        return await self.request_raw(
            session,
            "PUT",
            f"{_VTS}/vendor/TB_AE;{vendor}/calendars",
            json_body=calendar,
        )

    async def put_vendor_status(
        self, session: LoadedSession, vendor: str, status: str
    ) -> Any:
        """Partner Outlet Management holiday close / reopen.

        `PUT .../vendors/{id}/status` with `OPEN` / `CLOSED_TODAY` / `CLOSED_UNTIL`.
        Used for a closed weekday or a holiday; the next open day's hours write
        sends `OPEN` again.
        """
        return await self.request_json(
            session,
            "PUT",
            f"{_MENU_API}/vendors/{vendor}/status",
            json_body={"status": status},
        )

    async def prepare_session(
        self, db: AsyncSession, session: LoadedSession | None
    ) -> LoadedSession | None:
        """Fill store/finance account descriptors from `aggregator_branch_map`.

        The bootstrap often captures only the bearer and anti-bot cookies; the
        outlet ids the Report Builder and finance queries scope to live in the
        branch map (same pattern as Deliveroo's `prepare_session`).
        """
        if session is None:
            return None
        from app.services.aggregators import mapping

        tokens = dict(session.tokens or {})
        changed = False
        if not self._account_ids_in(tokens):
            outlets = await mapping.outlet_ids_for_channel(db, self.channel)
            if outlets:
                tokens["account_ids"] = outlets
                changed = True
        if not self._finance_accounts(replace(session, tokens=tokens)):
            accounts = await mapping.finance_accounts_for_channel(db, self.channel)
            if accounts:
                tokens["finance_accounts"] = accounts
                changed = True
        if changed:
            return replace(session, tokens=tokens)
        return session

    # ── GraphQL transport ─────────────────────────────────────────────────────
    async def _graphql(
        self,
        session: LoadedSession,
        *,
        endpoint: str,
        query: str,
        variables: dict[str, Any],
        operation_name: str,
    ) -> dict[str, Any]:
        """One GraphQL POST, returning its `data` object.

        Adds the bearer and the entity header, and turns a GraphQL-level
        `errors` array into `AggregatorUnavailableError` — the base already maps
        transport/auth failures, so only the application-level error is left to
        raise here.
        """
        headers = {
            "content-type": "application/json",
            "x-global-entity-id": self._global_entity_id(session).lower(),
        }
        token = self._access_token(session)
        if token:
            headers["authorization"] = f"Bearer {token}"
        payload = await self.request_json(
            session,
            "POST",
            endpoint,
            headers=headers,
            json_body={
                "query": query,
                "variables": variables,
                "operationName": operation_name,
            },
        )
        if isinstance(payload, dict) and payload.get("errors"):
            errors = payload["errors"]
            # A 200 whose errors carry an expired/invalid token is a dead session,
            # not a transient upstream error — raise the auth failure so the sweep
            # flips the session to `needs_bootstrap` (a re-login) and the run row
            # reads "session expired", instead of retrying a session that can only
            # be revived by a browser bootstrap.
            if _graphql_errors_are_auth(errors):
                raise AggregatorAuthError(
                    f"{self.channel} GraphQL {operation_name}: session token "
                    "expired — needs a re-login"
                )
            raise AggregatorUnavailableError(
                f"{self.channel} GraphQL {operation_name} returned errors: {errors}"
            )
        data = payload.get("data") if isinstance(payload, dict) else None
        if not isinstance(data, dict):
            raise AggregatorUnavailableError(
                f"{self.channel} GraphQL {operation_name} returned no data object"
            )
        return data

    # ── sales (Report Builder async export → CSV) ─────────────────────────────
    async def fetch_sales(
        self, session: LoadedSession, *, since: datetime, until: datetime
    ) -> SalesResult:
        account_ids = self._order_account_ids(session)
        if not account_ids:
            raise AggregatorUnavailableError(
                f"{self.channel} session carries no order store ids "
                "(tokens.account_ids / finance accounts) — cannot scope an export"
            )
        entity = self._global_entity_id(session)
        from_iso = since.date().isoformat()
        # Talabat's export `to` is EXCLUSIVE (a `to` of the 27th returns the 26th),
        # while `until` here is the LAST day to include — so advance `to` by one day.
        # This also keeps a single-day window (`until.date() == since.date()`, the
        # "yesterday" pull) valid: without it `from == to` is an empty range and the
        # portal rejects it with a VALIDATION_ERROR from vp-report-builder.
        to_iso = (until.date() + timedelta(days=1)).isoformat()
        filters = {
            "from": from_iso,
            "to": to_iso,
            "accountType": _EXPORT_ACCOUNT_TYPE_VENDOR,
            "accountIds": account_ids,
        }

        truncation_note: str | None = None
        estimate = await self._graphql(
            session,
            endpoint=_ORDERS_GRAPHQL,
            query=_ESTIMATE_EXPORT_SIZE_QUERY,
            variables={
                "input": {
                    "globalEntityId": entity,
                    "exportType": _EXPORT_TYPE_ORDERS,
                    "filters": filters,
                }
            },
            operation_name="EstimateExportSize",
        )
        size = estimate.get("estimateExportSize")
        if isinstance(size, dict) and size.get("withinLimits") is False:
            truncation_note = (
                "Talabat estimated the order export exceeds the portal's file "
                f"limits ({size}); the export may be capped short of the window."
            )

        requested = await self._graphql(
            session,
            endpoint=_ORDERS_GRAPHQL,
            query=_REQUEST_EXPORT_MUTATION,
            variables={
                "input": {
                    "name": f"sales-{from_iso}-{to_iso}",
                    "globalEntityId": entity,
                    "exportType": _EXPORT_TYPE_ORDERS,
                    "filters": filters,
                    "format": _EXPORT_FORMAT_CSV,
                    "deliveryMethod": _DELIVERY_METHOD_DIRECT,
                    "locale": "en",
                }
            },
            operation_name="RequestExport",
        )
        request_export = requested.get("requestExport")
        if not isinstance(request_export, dict) or not request_export.get("exportId"):
            raise AggregatorUnavailableError(
                f"{self.channel} RequestExport returned no export id: {requested}"
            )
        export_id = str(request_export["exportId"])

        ready = await self._poll_export_ready(session, export_id=export_id)
        download_url = str(ready.get("downloadUrl") or "")
        if not download_url:
            raise AggregatorUnavailableError(
                f"{self.channel} export {export_id} completed without a download URL"
            )
        csv_text = await self._download_csv(session, download_url)
        orders = self._orders_from_csv(csv_text)
        return SalesResult(orders=orders, truncation_note=truncation_note)

    async def _poll_export_ready(
        self, session: LoadedSession, *, export_id: str
    ) -> dict[str, Any]:
        """Poll `ListExports` until our export is completed (or terminally bad)."""
        entity = self._global_entity_id(session)
        deadline = asyncio.get_event_loop().time() + _EXPORT_POLL_TIMEOUT_SECONDS
        last: dict[str, Any] | None = None
        while asyncio.get_event_loop().time() < deadline:
            data = await self._graphql(
                session,
                endpoint=_ORDERS_GRAPHQL,
                query=_LIST_EXPORTS_QUERY,
                variables={
                    "input": {
                        "globalEntityId": entity,
                        "pagination": {"pageSize": 20, "sortOrder": "SORT_ORDER_DESC"},
                        "filter": {"exportType": []},
                    }
                },
                operation_name="ListExports",
            )
            exports = (data.get("listExports") or {}).get("exports") or []
            for export in exports:
                if not isinstance(export, dict) or export.get("exportId") != export_id:
                    continue
                last = export
                status = str(export.get("status") or "")
                if status == _EXPORT_STATUS_COMPLETED and export.get("downloadUrl"):
                    return export
                if status in _EXPORT_STATUS_TERMINAL_BAD:
                    raise AggregatorUnavailableError(
                        f"{self.channel} export {export_id} ended as {status}"
                    )
            await asyncio.sleep(_EXPORT_POLL_INTERVAL_SECONDS)
        raise AggregatorUnavailableError(
            f"{self.channel} export {export_id} did not finish in "
            f"{_EXPORT_POLL_TIMEOUT_SECONDS}s (last seen: {last})"
        )

    async def _download_csv(self, session: LoadedSession, download_url: str) -> str:
        """GET the completed export's CSV, replaying the bearer as the browser did."""
        token = self._access_token(session)
        headers = {"authorization": f"Bearer {token}"} if token else None
        response = await self.request_raw(
            session,
            "GET",
            download_url,
            headers=headers,
            timeout=_CSV_DOWNLOAD_TIMEOUT_SECONDS,
        )
        if self._is_auth_failure(response):
            # A 401/403 or a PerimeterX challenge on the download is a dead
            # session, not slowness — flip needs_bootstrap rather than loop.
            raise AggregatorAuthError(
                f"{self.channel} report download was challenged/blocked "
                "— session no longer authenticates"
            )
        status = getattr(response, "status_code", 0)
        if status >= 400:
            raise AggregatorUnavailableError(
                f"{self.channel} report download returned HTTP {status}"
            )
        content = getattr(response, "content", None)
        if isinstance(content, (bytes, bytearray)):
            return bytes(content).decode("utf-8-sig", errors="replace")
        return str(getattr(response, "text", "") or "")

    def _orders_from_csv(self, csv_text: str) -> list[StandardOrder]:
        """Every order row in the export CSV, with its parsed line items.

        Columns and money mapping are ported from the exporter's
        `parse_talabat_order_details_csv` / `parse_talabat_order_items_csv`. The
        outlet is left as Talabat's own `Store ID` (the ingest resolves it to a
        branch via `aggregator_branch_map`), and — unlike the scraper, which
        dropped everything but `delivered` — all statuses are kept, verbatim, so
        cancellations remain in the sales truth for reconciliation.
        """
        reader = csv.DictReader(io.StringIO(csv_text))
        orders: list[StandardOrder] = []
        for row in reader:
            external = (row.get("Order ID") or "").strip()
            if not external:
                continue
            placed_at = _parse_dt(row.get("Order received at"))
            status = (row.get("Order status") or "").strip() or None
            subtotal = _money(row.get("Subtotal"))
            orders.append(
                StandardOrder(
                    external_order_id=external,
                    external_outlet_id=(row.get("Store ID") or "").strip() or None,
                    business_date=placed_at.date().isoformat() if placed_at else None,
                    placed_at=placed_at,
                    status=status,
                    status_events=_status_events_from_row(row),
                    currency="AED",
                    gross_sales=subtotal,
                    net_sales=subtotal,
                    commission_amount=_money(row.get("Commission")),
                    payment_fee=_money(row.get("Online Payment Fee")),
                    # Talabat direct delivery: the vendor CSV carries no delivery
                    # fee column at all, so it stays unknown (None) rather than 0.
                    delivery_fee=None,
                    vat_amount=_money(_first(row, "Tax Amount", "Tax Charge")),
                    cancellation_fee=_money(row.get("Avoidable cancellation fee")),
                    net_payable=_money(
                        _first(row, "Payout Amount", "Estimated earnings")
                    ),
                    items=self._items_from_row(row, external, subtotal),
                    raw=dict(row),
                )
            )
        return orders

    @staticmethod
    def _items_from_row(
        row: dict[str, Any], external_order_id: str, subtotal: Decimal | None
    ) -> list[StandardOrderItem]:
        """Line items parsed out of the CSV's free-text `Order Items` cell.

        The export gives per-line money only implicitly: when a row has exactly
        one item, that item's amount is the order subtotal; with several items
        the per-line split is unknown, so `amount_is_known` is False and the
        money is left null rather than guessed.

        Modifiers are extracted from the item name token via two patterns:
        - Trailing parenthetical: `"Burger (No pickle, Extra cheese)"`
        - Plus-addon chain:       `"Burger + Extra sauce"`

        When the Report Builder CSV carries a dedicated modifier column
        (`Modifier names` / `Modifiers` / `Add-ons`) and there is exactly one
        item in the order (so the column's text unambiguously belongs to it),
        that column takes priority over the free-text extraction.
        """
        parsed = _parse_items_text(row.get("Order Items"))
        known = len(parsed) == 1
        # Dedicated modifier column — only reliable when there's one item.
        col_mod_text: str | None = None
        if known:
            raw_col = _first(row, "Modifier names", "Modifiers", "Add-ons")
            if isinstance(raw_col, str) and raw_col.strip():
                col_mod_text = raw_col.strip()

        items: list[StandardOrderItem] = []
        for index, (quantity, raw_name) in enumerate(parsed, start=1):
            # Extract modifiers from the name token (parenthetical / + chain).
            name, free_mod_texts = _extract_item_modifiers(raw_name)
            if col_mod_text and known:
                # Column takes priority; split on comma/semicolon.
                mod_texts = [
                    s.strip() for s in re.split(r"[,;]+", col_mod_text) if s.strip()
                ]
            else:
                mod_texts = free_mod_texts

            modifiers = [
                StandardModifier(name=m, quantity=Decimal("1")) for m in mod_texts if m
            ]
            modifiers_text = raw_name if modifiers else None

            unit_price: Decimal | None = None
            if known and subtotal is not None and quantity:
                unit_price = subtotal / quantity
            items.append(
                StandardOrderItem(
                    source_key=f"{external_order_id}:{index}",
                    grain=GRAIN_LINE,
                    item_name=name,
                    quantity=quantity,
                    unit_price=unit_price,
                    gross_sales=subtotal if known else None,
                    net_sales=subtotal if known else None,
                    amount_is_known=known,
                    modifiers=modifiers,
                    modifiers_text=modifiers_text,
                )
            )
        return items

    # ── finance (statements + payouts as distinct surfaces) ───────────────────
    async def fetch_statements(
        self, session: LoadedSession, *, since: datetime, until: datetime
    ) -> StatementsResult:
        accounts = self._finance_accounts(session)
        if not accounts:
            raise AggregatorUnavailableError(
                f"{self.channel} session carries no finance accounts "
                "(tokens.finance_accounts) — cannot query statements"
            )
        from_date = since.date()
        to_date = until.date()
        statement_rows = await self._paginate_finance(
            session,
            accounts=accounts,
            from_date=from_date,
            to_date=to_date,
            query=_LIST_ADDITIONAL_STATEMENTS_QUERY,
            operation_name="ListAdditionalStatements",
            root_key="listAdditionalStatements",
            item_key="additionalStatements",
            date_keys=("statementDateFrom", "statementDateTo"),
        )
        statements: list[StandardStatement] = [
            s
            for s in (self._statement_from(r) for r in statement_rows)
            if s is not None
        ]
        # Attempt to enrich with the detailed xlsx bundle (per-order fee lines).
        # If the bulk download fails — a network blip, a missing permission, a
        # provider 5xx — we degrade gracefully: the metadata statements above are
        # still correct, but the per-order lines/fees are absent, so we flag that
        # in `truncation_note` rather than letting a metadata-only result pass for
        # a complete one. (A dead session raises AggregatorAuthError, which is not
        # caught here so it can flip needs_bootstrap.)
        truncation_note: str | None = None
        try:
            bundle_statements = await self._fetch_bundle_statements(
                session, accounts=accounts, from_date=from_date, to_date=to_date
            )
            statements.extend(bundle_statements)
        except AggregatorUnavailableError as exc:
            logger.warning(
                "%s bundle download skipped (%s); metadata statements only",
                self.channel,
                exc,
            )
            truncation_note = (
                "Talabat detailed statement bundle (xlsx) could not be "
                f"downloaded ({exc}); per-order statement lines and fee/VAT "
                "breakdowns are absent — the statements returned carry "
                "settlement metadata only."
            )
        # The `ListAdditionalStatements` feed is only Talabat's *adjustment*
        # statements and is usually empty; the primary settlement metadata lives
        # inside each payout's `invoices[]` (invoiceId/period/totalPayout). Derive a
        # statement per invoice so every payout has a statement to couple to — the
        # payments leg of the chain — deduped by id so a real additional-statement
        # or bundle row always wins over the derived one.
        seen = {s.statement_id for s in statements}
        for invoice in await self._payout_invoice_rows(
            session, accounts=accounts, from_date=from_date, to_date=to_date
        ):
            if invoice.statement_id not in seen:
                statements.append(invoice)
                seen.add(invoice.statement_id)
        statements = await self._enrich_statements_with_attachment_lines(
            session, statements
        )
        if (
            truncation_note is None
            and statements
            and not any(s.lines for s in statements)
        ):
            truncation_note = (
                "Talabat statements carry no per-order lines — the xlsx bundle "
                "and per-statement attachments were empty or unreadable."
            )
        return StatementsResult(statements=statements, truncation_note=truncation_note)

    async def _payout_invoice_rows(
        self,
        session: LoadedSession,
        *,
        accounts: list[dict[str, Any]],
        from_date: date,
        to_date: date,
    ) -> list[StandardStatement]:
        """A statement per invoice inside the `ListPayouts` rows.

        Each payout batches one or more invoices, and the invoice sub-object is the
        only place Talabat exposes primary settlement metadata over httpx. The
        derived statement's `payment_due_date` is the payout's own transfer date, so
        `link_statements_to_payouts`'s date roll-up couples it back to that payout.
        """
        rows = await self._paginate_finance(
            session,
            accounts=accounts,
            from_date=from_date,
            to_date=to_date,
            query=_LIST_PAYOUTS_QUERY,
            operation_name="ListPayouts",
            root_key="listPayouts",
            item_key="payouts",
            date_keys=("startDate", "endDate"),
        )
        out: list[StandardStatement] = []
        for row in rows:
            transfer_date = _date_str(_first(row, "at", "paymentDateLocal"))
            invoices = row.get("invoices")
            if not isinstance(invoices, list):
                continue
            for invoice in invoices:
                if not isinstance(invoice, dict):
                    continue
                invoice_id = str(invoice.get("invoiceId") or "").strip()
                if not invoice_id:
                    continue
                period = invoice.get("period") if isinstance(invoice, dict) else None
                period = period if isinstance(period, dict) else {}
                amount = _money(_first(invoice, "invoiceAmount", "totalPayout"))
                out.append(
                    StandardStatement(
                        statement_id=invoice_id,
                        period_start=_date_str(period.get("from")),
                        period_end=_date_str(period.get("to")),
                        payment_due_date=_date_str(invoice.get("processedDate"))
                        or transfer_date,
                        gross_sales=amount,
                        net_payable=amount,
                        currency=str(
                            _first(invoice, "invoiceCurrency", "currency") or "AED"
                        ),
                        raw=dict(invoice),
                    )
                )
        return out

    async def fetch_payouts(
        self, session: LoadedSession, *, since: datetime, until: datetime
    ) -> PayoutsResult:
        accounts = self._finance_accounts(session)
        if not accounts:
            raise AggregatorUnavailableError(
                f"{self.channel} session carries no finance accounts "
                "(tokens.finance_accounts) — cannot query payouts"
            )
        from_date = since.date()
        to_date = until.date()
        payout_rows = await self._paginate_finance(
            session,
            accounts=accounts,
            from_date=from_date,
            to_date=to_date,
            query=_LIST_PAYOUTS_QUERY,
            operation_name="ListPayouts",
            root_key="listPayouts",
            item_key="payouts",
            date_keys=("startDate", "endDate"),
        )
        payouts = [self._payout_from(r) for r in payout_rows]
        payouts = [p for p in payouts if p is not None]
        return PayoutsResult(payouts=payouts)

    async def _paginate_finance(
        self,
        session: LoadedSession,
        *,
        accounts: list[dict[str, Any]],
        from_date: date,
        to_date: date,
        query: str,
        operation_name: str,
        root_key: str,
        item_key: str,
        date_keys: tuple[str, str],
    ) -> list[dict[str, Any]]:
        """Walk a finance listing to exhaustion, deduping and token-loop-guarded.

        Ported from `_talabat_paginate_finance_rows`: it stops on an absent
        `nextPageToken`, on a token it has already followed, and on the hard page
        ceiling — any one of which prevents an infinite live-portal loop.
        """
        entity = self._global_entity_id(session)
        rows: list[dict[str, Any]] = []
        seen_ids: set[str] = set()
        seen_tokens: set[str] = set()
        page_token: str | None = None
        for _ in range(_MAX_FINANCE_PAGES):
            pagination: dict[str, Any] = {"pageSize": _FINANCE_PAGE_SIZE}
            if page_token:
                pagination["pageToken"] = page_token
            params = {
                "globalEntityId": entity,
                "accounts": accounts,
                date_keys[0]: from_date.isoformat(),
                date_keys[1]: to_date.isoformat(),
                "filter": {},
                "pagination": pagination,
            }
            data = await self._graphql(
                session,
                endpoint=_FINANCE_GRAPHQL,
                query=query,
                variables={"params": params},
                operation_name=operation_name,
            )
            listing = (data.get("finances") or {}).get(root_key)
            if not isinstance(listing, dict):
                break
            for item in listing.get(item_key) or []:
                if not isinstance(item, dict):
                    continue
                row_id = str(_first(item, "payoutId", "statementId", "id") or id(item))
                if row_id not in seen_ids:
                    seen_ids.add(row_id)
                    rows.append(item)
            page_token = str(listing.get("nextPageToken") or "").strip() or None
            if not page_token or page_token in seen_tokens:
                break
            seen_tokens.add(page_token)
        return rows

    # ── statement bundle download + parse ────────────────────────────────────
    async def _fetch_bundle_statements(
        self,
        session: LoadedSession,
        *,
        accounts: list[dict[str, Any]],
        from_date: date,
        to_date: date,
    ) -> list[StandardStatement]:
        """Download the Talabat bulk statement zip and parse per-branch xlsx sheets.

        The portal's `bulkDownloadAdditionalStatements` GraphQL returns a
        pre-signed download URL for a zip that contains `Detailed_*.xlsx` files.
        httpx can fetch that URL with just the bearer token — no browser needed.
        If the file count exceeds the portal's direct-download limit the window
        is chunked into calendar months, matching the automation's behaviour.

        Raises `AggregatorUnavailableError` on any transport or API failure so
        the caller can degrade gracefully without dropping the metadata statements.
        """
        entity = self._global_entity_id(session)
        counts_data = await self._graphql(
            session,
            endpoint=_FINANCE_GRAPHQL,
            query=_BULK_STATEMENT_COUNTS_QUERY,
            variables={
                "params": {
                    "globalEntityId": entity,
                    "accounts": accounts,
                    "statementDateFrom": from_date.isoformat(),
                    "statementDateTo": to_date.isoformat(),
                    "filter": {},
                }
            },
            operation_name="GetBulkAdditionalStatementDownloadCounts",
        )
        counts_obj = (counts_data.get("finances") or {}).get(
            "getBulkAdditionalStatementDownloadCounts"
        ) or {}
        file_limits = counts_obj.get("fileLimits") or {}
        direct_limit = int(file_limits.get("directDownloadLimit") or 0)
        total_files = int(
            (counts_obj.get("fileCounts") or {}).get("totalFilesCount") or 0
        )
        windows = [(from_date, to_date)]
        if direct_limit and total_files > direct_limit:
            windows = _month_windows(from_date, to_date)
            logger.info(
                "%s chunking statement download into %d monthly windows "
                "(direct limit=%d, total=%d)",
                self.channel,
                len(windows),
                direct_limit,
                total_files,
            )

        all_statements: list[StandardStatement] = []
        for win_start, win_end in windows:
            url = await self._bundle_download_url(
                session, accounts=accounts, from_date=win_start, to_date=win_end
            )
            if not url:
                logger.debug(
                    "%s no bundle download URL for window %s – %s",
                    self.channel,
                    win_start,
                    win_end,
                )
                continue
            bundle_bytes = await self._download_bundle(session, url)
            parsed = self._parse_bundle_bytes(bundle_bytes)
            # Archive the downloaded zip (the VAT-claim source document) to R2 and
            # stamp the invoice fields onto every statement parsed out of it.
            parsed = self._archive_bundle(
                parsed, bundle_bytes, from_date=win_start, to_date=win_end
            )
            logger.info(
                "%s bundle %s – %s: %d detailed statements",
                self.channel,
                win_start,
                win_end,
                len(parsed),
            )
            all_statements.extend(parsed)
        return all_statements

    def _archive_bundle(
        self,
        statements: list[StandardStatement],
        bundle_bytes: bytes,
        *,
        from_date: date,
        to_date: date,
    ) -> list[StandardStatement]:
        """Persist the statement zip to R2 and stamp the invoice fields on it.

        The zip is one physical document — the portal's own
        `additionalStatementsArchive_…zip` — covering every branch statement in
        the window, so each parsed statement points at the same archived object.

        `store_statement_invoice` returns None when R2 is unconfigured or the
        body is empty; then the statements are returned unchanged, still
        carrying their parsed totals and lines (archival is best-effort — it must
        never drop a statement). The same is true if the upload itself throws.
        """
        if not statements or not bundle_bytes:
            return statements
        from app.services.aggregators import statement_docs

        filename = (
            f"additionalStatementsArchive_{from_date.isoformat()}"
            f"_to_{to_date.isoformat()}.zip"
        )
        archive_id = f"bundle-{from_date.isoformat()}-{to_date.isoformat()}"
        try:
            stored = statement_docs.store_statement_invoice(
                channel=self.channel,
                statement_id=archive_id,
                filename=filename,
                body=bundle_bytes,
                content_type="application/zip",
            )
        except Exception:  # noqa: BLE001 - an archive failure must not drop lines
            logger.warning(
                "%s statement bundle archive failed for %s – %s; "
                "statements kept without an invoice document",
                self.channel,
                from_date,
                to_date,
                exc_info=True,
            )
            return statements
        if stored is None:
            return statements
        return [
            replace(
                stmt,
                invoice_object_key=stored.object_key,
                invoice_content_type=stored.content_type,
                invoice_original_filename=stored.original_filename,
                invoice_fetched_at=stored.fetched_at,
                invoice_attachments=stored.attachments,
            )
            for stmt in statements
        ]

    async def _bundle_download_url(
        self,
        session: LoadedSession,
        *,
        accounts: list[dict[str, Any]],
        from_date: date,
        to_date: date,
    ) -> str | None:
        """Request the bulk statement zip download URL for a single date window."""
        entity = self._global_entity_id(session)
        data = await self._graphql(
            session,
            endpoint=_FINANCE_GRAPHQL,
            query=_BULK_STATEMENT_DOWNLOAD_QUERY,
            variables={
                "params": {
                    "globalEntityId": entity,
                    "accounts": accounts,
                    "statementDateFrom": from_date.isoformat(),
                    "statementDateTo": to_date.isoformat(),
                    "filter": {},
                }
            },
            operation_name="RequestBulkDownloadAdditionalStatements",
        )
        url = str(
            (data.get("finances") or {})
            .get("bulkDownloadAdditionalStatements", {})
            .get("downloadUrl")
            or ""
        ).strip()
        return self._resolve_download_url(url) if url else None

    async def _download_bundle(
        self, session: LoadedSession, download_url: str
    ) -> bytes:
        """GET the bundle zip bytes, replaying the bearer as the browser did."""
        token = self._access_token(session)
        headers = {"authorization": f"Bearer {token}"} if token else None
        response = await self.request_raw(
            session,
            "GET",
            self._resolve_download_url(download_url),
            headers=headers,
            timeout=_CSV_DOWNLOAD_TIMEOUT_SECONDS,
        )
        if self._is_auth_failure(response):
            # A 401/403 or a PerimeterX challenge on the download is a dead
            # session, not slowness — flip needs_bootstrap rather than loop.
            raise AggregatorAuthError(
                f"{self.channel} bundle download was challenged/blocked "
                "— session no longer authenticates"
            )
        status = getattr(response, "status_code", 0)
        if status >= 400:
            raise AggregatorUnavailableError(
                f"{self.channel} bundle download returned HTTP {status}"
            )
        content = getattr(response, "content", None)
        if isinstance(content, (bytes, bytearray)):
            return bytes(content)
        raise AggregatorUnavailableError(
            f"{self.channel} bundle download returned no binary content"
        )

    @staticmethod
    def _resolve_download_url(url: str) -> str:
        """Absolute URL for a finance/menu attachment.

        GraphQL sometimes returns a relative vendor-api path. The menu reads
        already use that host; a relative GET against the GraphQL origin 404s
        and is how statement lines stayed empty.
        """
        value = (url or "").strip()
        if not value:
            return value
        if value.startswith("//"):
            return f"https:{value}"
        if value.startswith("http://") or value.startswith("https://"):
            return value
        if value.startswith("/"):
            return f"{_VENDOR_API_ORIGIN}{value}"
        return f"{_VENDOR_API_ORIGIN}/{value}"

    @staticmethod
    def _attachment_urls(raw: dict[str, Any] | None) -> list[str]:
        """Download URLs/paths off a ListAdditional or invoice row."""
        if not isinstance(raw, dict):
            return []
        atts = (
            raw.get("attachments")
            or raw.get("invoiceAttachments")
            or raw.get("payoutAttachments")
            or []
        )
        urls: list[str] = []
        if isinstance(atts, str) and atts.strip():
            return [atts.strip()]
        if not isinstance(atts, list):
            return []
        for item in atts:
            if isinstance(item, str) and item.strip():
                urls.append(item.strip())
            elif isinstance(item, dict):
                path = item.get("path") or item.get("url") or item.get("downloadUrl")
                if path:
                    urls.append(str(path).strip())
        return urls

    @staticmethod
    def _as_bundle_zip(body: bytes) -> bytes:
        """A zip the existing Detailed_*.xlsx parser accepts.

        Per-statement attachments are often a lone xlsx; the bulk download is
        already a zip. Wrap the former so `_parse_bundle_bytes` is one parser.
        """
        if body[:2] == b"PK":
            return body
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as archive:
            archive.writestr("Detailed_attachment.xlsx", body)
        return buf.getvalue()

    async def _enrich_statements_with_attachment_lines(
        self, session: LoadedSession, statements: list[StandardStatement]
    ) -> list[StandardStatement]:
        """Fill empty `lines` from each statement's own xlsx attachment.

        Idempotent: statements that already have lines (the bulk bundle) are
        left alone. A failed attachment is skipped so one bad file cannot drop
        the metadata row.
        """
        out: list[StandardStatement] = []
        for stmt in statements:
            if stmt.lines:
                out.append(stmt)
                continue
            lines: list[StandardStatementLine] = []
            for url in self._attachment_urls(stmt.raw):
                try:
                    body = await self._download_bundle(session, url)
                    parsed = self._parse_bundle_bytes(self._as_bundle_zip(body))
                except AggregatorAuthError:
                    raise
                except AggregatorUnavailableError as exc:
                    logger.warning(
                        "%s attachment %s for %s skipped (%s)",
                        self.channel,
                        url,
                        stmt.statement_id,
                        exc,
                    )
                    continue
                for parsed_stmt in parsed:
                    for line in parsed_stmt.lines:
                        lines.append(
                            replace(
                                line,
                                statement_id=stmt.statement_id,
                                source_key=(f"{stmt.statement_id}:{line.source_key}"),
                            )
                        )
            out.append(replace(stmt, lines=lines) if lines else stmt)
        return out

    @staticmethod
    def _parse_bundle_bytes(bundle_bytes: bytes) -> list[StandardStatement]:
        """Parse a Talabat statement zip bundle into per-branch StandardStatement objects.

        Each `Detailed_*.xlsx` inside the zip covers one billing period. Rows are
        per-order; we accumulate totals and per-order fee lines per branch, then
        emit one `StandardStatement` per branch with those lines attached.

        The statement_id uses `detailed-{start}-{end}-{branch}` — the same
        per-branch key the automation uses — so two branches in one period produce
        two distinct statements rather than one overwriting the other (see the
        automation comments on the unique-key bug that was fixed there).
        """
        statements: list[StandardStatement] = []
        try:
            archive = zipfile.ZipFile(io.BytesIO(bundle_bytes))
        except zipfile.BadZipFile as exc:
            raise AggregatorUnavailableError(
                f"Talabat bundle is not a valid zip: {exc}"
            ) from exc

        with archive:
            workbook_names = [
                name
                for name in archive.namelist()
                if name.lower().endswith(".xlsx") and "detailed_" in name.lower()
            ]
            for member_name in workbook_names:
                wb = load_workbook(
                    io.BytesIO(archive.read(member_name)),
                    data_only=True,
                    read_only=True,
                )
                sheet = wb[wb.sheetnames[0]]
                period_start, period_end = _bundle_period(sheet["A1"].value)

                header_map: dict[str, int] = {}
                header_row_idx: int | None = None
                for row_idx, row in enumerate(
                    sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1
                ):
                    cands = _bundle_headers(row)
                    if "Order Id" in cands and "Net Payment Per Order" in cands:
                        header_row_idx = row_idx
                        header_map = cands
                        break
                if header_row_idx is None:
                    logger.debug(
                        "Talabat bundle sheet %s: no recognisable header row",
                        member_name,
                    )
                    continue

                start_token = (
                    period_start.isoformat() if period_start else "unknown-start"
                )
                end_token = period_end.isoformat() if period_end else "unknown-end"

                # per_branch_totals[branch_id] = {gross_sales, net_payable, total_fees, total_vat}
                per_branch_totals: dict[str, dict[str, Decimal]] = {}
                per_branch_lines: dict[str, list[StandardStatementLine]] = {}

                for row in sheet.iter_rows(
                    min_row=header_row_idx + 1, values_only=True
                ):
                    order_id_raw = _bundle_cell(row, header_map, "Order Id")
                    if not _bundle_is_order_id(order_id_raw):
                        continue
                    external_order_id = _bundle_order_id_str(order_id_raw)
                    branch_id = str(
                        _bundle_cell(row, header_map, "Branch Id") or ""
                    ).strip()
                    branch_token = f"-{branch_id}" if branch_id else ""
                    statement_id = f"detailed-{start_token}-{end_token}{branch_token}"

                    # Parse line date
                    line_date_raw = _bundle_cell(row, header_map, "Date / Time")
                    line_date_str: str | None = None
                    if isinstance(line_date_raw, datetime):
                        line_date_str = line_date_raw.date().isoformat()
                    elif line_date_raw:
                        d = _parse_date(str(line_date_raw).split(" ")[0])
                        line_date_str = d.isoformat() if d else None

                    # Money helpers for this row
                    def m(col: str) -> Decimal:
                        return _bundle_money(row, header_map, col)

                    gross = m("SubTotal")
                    fee = (
                        m("Commission VAT Exclu.")
                        + m("Payment Handling Charges")
                        + m("Promotional Fees")
                        + m("Sponsored Deal Fees")
                        + m("Avoidable Wait Time Fee")
                        + m("Cost Per Order")
                        + m("GEM Fee")
                        + m("Loyalty Charges")
                    )
                    vat = (
                        m("Commission VAT")
                        + m("Payment Handling Charges VAT")
                        + m("Avoidable Wait Time Fee VAT")
                    )
                    net = m("Net Payment Per Order")

                    totals = per_branch_totals.setdefault(
                        branch_id,
                        {
                            "gross_sales": Decimal("0"),
                            "net_payable": Decimal("0"),
                            "total_fees": Decimal("0"),
                            "total_vat": Decimal("0"),
                        },
                    )
                    totals["gross_sales"] += gross
                    totals["total_fees"] += fee
                    totals["total_vat"] += vat
                    totals["net_payable"] += net

                    # Emit per-order fee lines (skip zero amounts)
                    lines_for_branch = per_branch_lines.setdefault(branch_id, [])
                    line_specs = [
                        ("gross_sales", "subtotal", gross),
                        ("fee", "commission", -abs(m("Commission VAT Exclu."))),
                        ("vat", "commission_vat", -abs(m("Commission VAT"))),
                        (
                            "fee",
                            "payment_handling",
                            -abs(m("Payment Handling Charges")),
                        ),
                        (
                            "vat",
                            "payment_handling_vat",
                            -abs(m("Payment Handling Charges VAT")),
                        ),
                        ("fee", "promotional_fees", -abs(m("Promotional Fees"))),
                        ("fee", "sponsored_deal_fees", -abs(m("Sponsored Deal Fees"))),
                        (
                            "fee",
                            "avoidable_wait_time_fee",
                            -abs(m("Avoidable Wait Time Fee")),
                        ),
                        (
                            "vat",
                            "avoidable_wait_time_fee_vat",
                            -abs(m("Avoidable Wait Time Fee VAT")),
                        ),
                        ("fee", "cost_per_order", -abs(m("Cost Per Order"))),
                        ("fee", "gem_fee", -abs(m("GEM Fee"))),
                        ("fee", "loyalty_charges", -abs(m("Loyalty Charges"))),
                        ("net_payable", "net_payable", net),
                    ]
                    for line_type, fee_category, amount in line_specs:
                        if amount == Decimal("0"):
                            continue
                        lines_for_branch.append(
                            StandardStatementLine(
                                source_key=f"{statement_id}:{external_order_id}:{fee_category}",
                                statement_id=statement_id,
                                external_order_id=external_order_id,
                                line_date=line_date_str,
                                line_type=line_type,
                                fee_category=fee_category,
                                description=(
                                    f"Talabat detailed statement {fee_category.replace('_', ' ')}"
                                ),
                                amount=amount,
                                currency="AED",
                            )
                        )

                for branch_id, totals in sorted(per_branch_totals.items()):
                    if totals["gross_sales"] == Decimal("0") and totals[
                        "net_payable"
                    ] == Decimal("0"):
                        continue
                    branch_token = f"-{branch_id}" if branch_id else ""
                    statement_id = f"detailed-{start_token}-{end_token}{branch_token}"
                    q = Decimal("0.01")
                    statements.append(
                        StandardStatement(
                            statement_id=statement_id,
                            period_start=period_start.isoformat()
                            if period_start
                            else None,
                            period_end=period_end.isoformat() if period_end else None,
                            payment_due_date=period_end.isoformat()
                            if period_end
                            else None,
                            gross_sales=totals["gross_sales"].quantize(q),
                            net_payable=totals["net_payable"].quantize(q),
                            total_fees=totals["total_fees"].quantize(q),
                            total_vat=totals["total_vat"].quantize(q),
                            currency="AED",
                            external_outlet_id=branch_id if branch_id else None,
                            lines=per_branch_lines.get(branch_id, []),
                        )
                    )

        return statements

    @staticmethod
    def _payout_from(row: dict[str, Any]) -> StandardPayout | None:
        """A `ListPayouts` row mapped to a payout — port of
        `normalize_talabat_graphql_payout_rows`."""
        payout_id = str(_first(row, "payoutId", "id") or "").strip()
        transfer_date = _date_str(_first(row, "at", "paymentDateLocal"))
        if not payout_id or not transfer_date:
            return None
        payment_reference = payout_id
        invoices = row.get("invoices")
        if isinstance(invoices, list) and invoices and isinstance(invoices[0], dict):
            invoice_id = str(invoices[0].get("invoiceId") or "").strip()
            if invoice_id:
                payment_reference = invoice_id
        status = str(_first(row, "status", "payoutStatus") or "").lower() or None
        return StandardPayout(
            transfer_id=payout_id,
            transfer_date=transfer_date,
            payment_due_date=transfer_date,
            transfer_amount=_money(_first(row, "payoutAmount", "netPayout")),
            transfer_status=status,
            payment_reference=payment_reference,
            currency=str(_first(row, "payoutCurrency", "currency") or "AED"),
        )

    @staticmethod
    def _statement_from(row: dict[str, Any]) -> StandardStatement | None:
        """A `ListAdditionalStatements` row mapped to a statement — port of
        `normalize_talabat_graphql_statement_rows`.

        This is the settlement *metadata* (id, date, gross). Per-order fee
        lines come from the row's `attachments` and/or the bulk xlsx bundle,
        both wired in `fetch_statements` via `_parse_bundle_bytes`.
        """
        statement_id = str(row.get("statementId") or "").strip()
        statement_date = _date_str(row.get("statementDate"))
        if not statement_id or not statement_date:
            return None
        gross = _money(row.get("amountGross"))
        return StandardStatement(
            statement_id=statement_id,
            period_end=statement_date,
            payment_due_date=statement_date,
            gross_sales=gross,
            net_payable=gross,
            currency=str(row.get("currency") or "AED"),
            raw=dict(row),
        )


#: The module-level singleton, matching the careem/grubops/foodics providers —
#: stateless (the session is passed per call), so sharing it is free.
provider = TalabatClient()
