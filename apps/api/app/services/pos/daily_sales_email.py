"""The daily sales email: one spreadsheet, every branch, every channel.

A branch owner does not open the manager app each morning; a mail with the
numbers already in it is read. This builds that mail — an `.xlsx` with five tabs,
all for the one frozen trading day: a **Summary** matrix whose five stacked
blocks (Sales Revenue, Order Count, Sales Discount, Charges, Net Revenue) each
carry one row per branch and one column per channel, plus per-row/column/grand
totals; and a tab each of the records behind it — **Orders** (order level),
**Statements** and **Statement Lines** (the marketplace settlement detail) and
**Payouts** — and sends it after the last branch has closed for the day.

Money basis, decided with the owner: **what the customer paid, VAT included.**
Sales Revenue is the order total; Net Revenue is that less our costs and any
refund — the same figure `order_economics.net` computes per order. Charges is
those costs summed with no breakdown: the marketplace's commission, the payment
fee, and — on a website order — what the courier cost us. Discount is shown for
information, since it is already inside the total.

Only **delivered** trade counts: a delivered marketplace or website order, or a
closed counter check. Anything cancelled or still in progress is left out — a
sales report is a record of money taken, not orders opened.

The loop belongs to the app because this stack has no cron, and it holds an
advisory lock so a second copy achieves nothing — the same shape as
`log_retention` and the batch dispatcher next to it in the lifespan.
"""

from __future__ import annotations

import asyncio
import logging
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import advisory_lock, trading_hours
from app.core.database import AsyncSessionFactory
from app.models.aggregator import (
    AggregatorPayout,
    AggregatorStatement,
    AggregatorStatementLine,
)
from app.models.branch import Branch
from app.models.email_log import EmailLog
from app.models.order import Order, OrderStatusEnum
from app.models.order_delivery import OrderDelivery
from app.models.pos_order import PosOrderStatusEnum
from app.services import email_service
from app.services.couriers import courier_catalog
from app.services.pos import business_day_service

logger = logging.getLogger(__name__)

_ZERO = Decimal("0")

#: Same flat 64-bit namespace as every other advisory lock. "mmBATCH" + 5.
_ADVISORY_LOCK_KEY = 0x6D6D_4241_5443_4805

#: How long after the last branch's configured close the mail goes out. The
#: report counts *delivered* trade, so a rider still out at close — a website or
#: aggregator order handed over after we snapshot — would be undercounted. Forty
#: five minutes lets that tail settle without pushing the mail deep into the
#: night; shorten it for a faster but rougher send.
_CLOSE_BUFFER = timedelta(minutes=45)

_TICK_SECONDS = 600

#: Where the automatic send goes. The manual admin trigger passes its own list.
DEFAULT_RECIPIENTS = ["h_abbasi97@hotmail.com", "fahimakhtarabbasi@gmail.com"]

_TEMPLATE = "daily_sales_report"

#: The channel columns, in report order. The five aggregators are the courier
#: catalogue's own codes; website folds every courier into one column and
#: counter is the till. A marketplace nobody has mapped yet still gets a column
#: (appended) rather than having its money silently dropped.
_AGGREGATOR_COLUMNS = ["keeta", "noon_food", "talabat", "careem", "deliveroo"]
_FIXED_COLUMNS = _AGGREGATOR_COLUMNS + ["website", "counter"]
_COLUMN_LABELS = {
    "keeta": "keeta",
    "noon_food": "noon food",
    "talabat": "talabat",
    "careem": "careem",
    "deliveroo": "deliveroo",
    "website": "website",
    "counter": "counter",
}


def _label(column: str) -> str:
    return _COLUMN_LABELS.get(column, column)


#: Delivered trade only. A counter check is done when the till closes it; a
#: marketplace or website order when it is delivered. Cancelled, refunded and
#: in-progress orders never match either arm.
_DELIVERED = or_(
    and_(
        Order.source == "cashier", Order.pos_status == PosOrderStatusEnum.CLOSED.value
    ),
    and_(
        Order.source.in_(["aggregator", "online"]),
        Order.status == OrderStatusEnum.DELIVERED.value,
    ),
)


@dataclass
class Cell:
    """One (branch, channel) intersection, before the sections read it."""

    revenue: Decimal = _ZERO  # order total, VAT included
    discount: Decimal = _ZERO
    charges: Decimal = _ZERO  # commission + payment fee + courier cost, our costs
    refunds: Decimal = _ZERO
    count: int = 0

    @property
    def net(self) -> Decimal:
        return self.revenue - self.charges - self.refunds


@dataclass
class ReportRow:
    business_date: str
    branch_name: str
    cells: dict[str, Cell]


@dataclass
class DailySalesReport:
    date_from: str
    date_to: str
    columns: list[str]
    rows: list[ReportRow] = field(default_factory=list)


#: Each block in the sheet, and how it reads a cell. Discount and Charges are
#: shown negative, the way the owner's mock-up does — money leaving the till.
_SECTIONS: list[tuple[str, "callable[[Cell], Decimal | int]"]] = [
    ("Sales Revenue", lambda c: c.revenue),
    ("Order Count", lambda c: c.count),
    ("Sales Discount", lambda c: -c.discount),
    ("Charges", lambda c: -c.charges),
    ("Net Revenue", lambda c: c.net),
]


def _column_for(source: str | None, aggregator_channel: str | None) -> str | None:
    if source == "cashier":
        return "counter"
    if source == "online":
        return "website"
    if source == "aggregator":
        code = courier_catalog.code_for_channel(aggregator_channel or "")
        return code or (aggregator_channel or "unknown")
    return None


async def _fetch(
    db: AsyncSession, date_from: str, date_to: str, branch_id=None
) -> list:
    # `cost_total` is what the courier actually billed; `quoted_cost` stands in
    # until it lands — the same order `order_economics` prefers them in. The
    # delivery join cannot fan out: `order_deliveries` is unique on order_id.
    courier_cost = func.coalesce(OrderDelivery.cost_total, OrderDelivery.quoted_cost)
    stmt = (
        select(
            Order.business_date,
            Order.branch_id,
            Order.source,
            Order.aggregator_channel,
            func.count(Order.id),
            func.coalesce(func.sum(Order.total), 0),
            func.coalesce(func.sum(Order.discount_amount), 0),
            func.coalesce(func.sum(func.coalesce(Order.aggregator_fee, 0)), 0),
            func.coalesce(func.sum(func.coalesce(Order.payment_fee, 0)), 0),
            func.coalesce(func.sum(func.coalesce(courier_cost, 0)), 0),
            func.coalesce(func.sum(func.coalesce(Order.refunded_amount, 0)), 0),
        )
        .select_from(Order)
        .outerjoin(OrderDelivery, OrderDelivery.order_id == Order.id)
        .where(Order.is_pos.is_(True))
        .where(Order.business_date >= date_from, Order.business_date <= date_to)
        .where(_DELIVERED)
        .group_by(
            Order.business_date, Order.branch_id, Order.source, Order.aggregator_channel
        )
    )
    if branch_id is not None:
        stmt = stmt.where(Order.branch_id == branch_id)
    return (await db.execute(stmt)).all()


async def build(
    db: AsyncSession, *, date_from: str, date_to: str, branch_id=None
) -> DailySalesReport:
    """Assemble the matrix: one row per (trading day, branch), zero-filled."""
    raw = await _fetch(db, date_from, date_to, branch_id)

    branches = (
        (await db.execute(select(Branch).where(Branch.is_active.is_(True))))
        .scalars()
        .all()
    )
    if branch_id is not None:
        branches = [b for b in branches if b.id == branch_id]
    branch_name: dict = {b.id: b.name for b in branches}

    matrix: dict[tuple[str, object], dict[str, Cell]] = {}
    extra_columns: list[str] = []
    dates_present: set[str] = set()

    for (
        bdate,
        bid,
        source,
        channel,
        cnt,
        revenue,
        discount,
        agg_fee,
        pay_fee,
        cour_cost,
        refunds,
    ) in raw:
        col = _column_for(source, channel)
        if col is None:
            continue
        if col not in _FIXED_COLUMNS and col not in extra_columns:
            extra_columns.append(col)
        # A branch that traded but is no longer active still owns its money.
        branch_name.setdefault(bid, "(inactive branch)")
        dates_present.add(bdate)
        cell = matrix.setdefault((bdate, bid), {}).setdefault(col, Cell())
        cell.revenue += revenue
        cell.discount += discount
        cell.charges += agg_fee + pay_fee + cour_cost
        cell.refunds += refunds
        cell.count += int(cnt or 0)

    columns = _FIXED_COLUMNS + extra_columns

    # Days that traded, oldest first; the requested date itself if none did, so
    # a quiet day still produces a report that says so rather than an empty file.
    dates = sorted(dates_present) or [date_from]
    ordered_bids = sorted(branch_name, key=lambda b: branch_name[b].lower())

    rows: list[ReportRow] = []
    for d in dates:
        for bid in ordered_bids:
            cells = matrix.get((d, bid), {})
            rows.append(
                ReportRow(
                    business_date=d,
                    branch_name=branch_name[bid],
                    cells={c: cells.get(c, Cell()) for c in columns},
                )
            )

    return DailySalesReport(
        date_from=date_from, date_to=date_to, columns=columns, rows=rows
    )


@dataclass
class ReportDetail:
    """The per-record backing sheets, all scoped to the same frozen business date
    range as the summary matrix: one row per order, per statement, per statement
    line, per payout."""

    orders: list = field(default_factory=list)
    statements: list = field(default_factory=list)
    statement_lines: list = field(default_factory=list)
    payouts: list = field(default_factory=list)


async def build_detail(
    db: AsyncSession, *, date_from: str, date_to: str, branch_id=None
) -> ReportDetail:
    """Fetch the order/statement/line/payout rows for the detail sheets.

    Orders are the same delivered-trade set the summary counts, keyed on the
    frozen business date. Finance rows (statement lines, payouts) are the events
    DATED to that range — a line whose `line_date` or a payout whose
    `transfer_date` falls in it — and the statements those lines roll up to, so
    the finance tabs show what settled on the day rather than the whole ledger.
    """
    courier_cost = func.coalesce(OrderDelivery.cost_total, OrderDelivery.quoted_cost)
    order_stmt = (
        select(
            Order.business_date,
            Order.order_number,
            Branch.name,
            Order.source,
            Order.aggregator_channel,
            Order.status,
            Order.pos_status,
            Order.customer_name,
            Order.total,
            Order.vat_amount,
            func.coalesce(Order.aggregator_fee, 0),
            func.coalesce(Order.payment_fee, 0),
            func.coalesce(courier_cost, 0),
            func.coalesce(Order.refunded_amount, 0),
            # The marketplace's OWN reference for this order, so a row can be
            # tied back to what the aggregator portal shows. `external_reference`
            # is the canonical id the marketplace assigns (Careem's order id,
            # Deliveroo's UUID, Noon's ref, …); `aggregator_display_code` is the
            # SHORT human code where it differs (Deliveroo's `5254`, Keeta's
            # last-4). Appended (o[14], o[15]) so the earlier positions are stable.
            Order.external_reference,
            Order.aggregator_display_code,
        )
        .select_from(Order)
        .outerjoin(OrderDelivery, OrderDelivery.order_id == Order.id)
        .outerjoin(Branch, Branch.id == Order.branch_id)
        .where(Order.is_pos.is_(True))
        .where(Order.business_date >= date_from, Order.business_date <= date_to)
        .where(_DELIVERED)
        .order_by(Order.business_date, Branch.name, Order.aggregator_channel)
    )
    if branch_id is not None:
        order_stmt = order_stmt.where(Order.branch_id == branch_id)
    orders = (await db.execute(order_stmt)).all()

    line_stmt = (
        select(AggregatorStatementLine)
        .where(
            AggregatorStatementLine.line_date >= date_from,
            AggregatorStatementLine.line_date <= date_to,
        )
        .order_by(
            AggregatorStatementLine.channel,
            AggregatorStatementLine.line_date,
            AggregatorStatementLine.statement_id,
        )
    )
    lines = (await db.execute(line_stmt)).scalars().all()

    # The statements those in-range lines belong to, plus any statement whose own
    # period falls in the range (a summary-grain statement with no per-day line).
    stmt_ids = {ln.statement_id for ln in lines if ln.statement_id}
    stmt_stmt = (
        select(AggregatorStatement)
        .where(
            or_(
                AggregatorStatement.statement_id.in_(stmt_ids) if stmt_ids else False,
                and_(
                    AggregatorStatement.period_start <= date_to,
                    AggregatorStatement.period_end >= date_from,
                ),
            )
        )
        .order_by(AggregatorStatement.channel, AggregatorStatement.period_start)
    )
    statements = (await db.execute(stmt_stmt)).scalars().all()

    payout_stmt = (
        select(AggregatorPayout)
        .where(
            AggregatorPayout.transfer_date >= date_from,
            AggregatorPayout.transfer_date <= date_to,
        )
        .order_by(AggregatorPayout.channel, AggregatorPayout.transfer_date)
    )
    payouts = (await db.execute(payout_stmt)).scalars().all()

    return ReportDetail(
        orders=orders,
        statements=statements,
        statement_lines=lines,
        payouts=payouts,
    )


def _num(value: Decimal | int) -> float | int:
    if isinstance(value, int):
        return value
    return float(round(value, 2))


def _dec(value) -> float:
    return float(round(value or _ZERO, 2))


def _write_sheet(ws, headers: list[str], rows: list[list], bold) -> None:
    """A simple header + rows sheet, header bold, columns auto-ish widened."""
    for ci, text in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=text).font = bold
    for ri, row in enumerate(rows, start=2):
        for ci, value in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=value)
    for ci, text in enumerate(headers, start=1):
        width = max(
            10,
            min(
                40,
                max(
                    [len(str(text))]
                    + [len(str(r[ci - 1])) if ci - 1 < len(r) else 0 for r in rows]
                )
                + 2,
            ),
        )
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = width


def _summary_sheet(ws, report: DailySalesReport, bold) -> None:
    """The branch×channel matrix, five stacked sections, each with a Total column
    (row sum) and a Totals row (column sum) — the summations the summary owes."""
    r = 1
    for title, metric in _SECTIONS:
        ws.cell(row=r, column=1, value=title).font = bold
        r += 1
        header = ["date", "branch"] + [_label(c) for c in report.columns] + ["TOTAL"]
        for ci, text in enumerate(header, start=1):
            ws.cell(row=r, column=ci, value=text).font = bold
        r += 1
        col_totals = [0.0 for _ in report.columns]
        grand = 0.0
        for row in report.rows:
            ws.cell(row=r, column=1, value=row.business_date)
            ws.cell(row=r, column=2, value=row.branch_name)
            row_total = 0.0
            for ci, col in enumerate(report.columns):
                v = _num(metric(row.cells[col]))
                ws.cell(row=r, column=3 + ci, value=v)
                col_totals[ci] += float(v)
                row_total += float(v)
            ws.cell(row=r, column=3 + len(report.columns), value=round(row_total, 2))
            grand += row_total
            r += 1
        # Totals row.
        ws.cell(row=r, column=2, value="TOTAL").font = bold
        for ci, tv in enumerate(col_totals):
            ws.cell(row=r, column=3 + ci, value=round(tv, 2)).font = bold
        ws.cell(
            row=r, column=3 + len(report.columns), value=round(grand, 2)
        ).font = bold
        r += 2  # blank row between blocks

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16


def to_xlsx(report: DailySalesReport, detail: ReportDetail | None = None) -> bytes:
    """Render the report to a multi-sheet workbook.

    Summary (the branch×channel matrix with per-row/column/grand totals), then a
    tab each of the underlying records for the same frozen date range: Orders,
    Statements, Statement Lines, Payouts. `detail=None` renders Summary only (the
    old single-report call), so existing callers keep working."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    summary = wb.active
    summary.title = "Summary"
    _summary_sheet(summary, report, bold)

    if detail is not None:
        _write_sheet(
            wb.create_sheet("Orders"),
            [
                "date",
                "order #",
                "aggregator ref",
                "aggregator code",
                "branch",
                "channel",
                "status",
                "customer",
                "total",
                "vat",
                "commission",
                "payment fee",
                "courier cost",
                "refund",
                "net",
            ],
            [
                # positional: 0 date,1 order#,2 branch,3 source,4 channel,5 status,
                # 6 pos_status,7 customer,8 total,9 vat,10 commission,11 payfee,
                # 12 courier,13 refund,14 external_reference,15 display_code
                [
                    o[0],
                    o[1],
                    # Aggregator refs only — a website/counter order has none to
                    # reconcile against a marketplace portal.
                    (o[14] or "") if o[3] == "aggregator" else "",
                    (o[15] or "") if o[3] == "aggregator" else "",
                    o[2] or "",
                    _order_channel_label(o[3], o[4]),
                    (o[5] or o[6] or ""),
                    o[7] or "",
                    _dec(o[8]),
                    _dec(o[9]),
                    _dec(o[10]),
                    _dec(o[11]),
                    _dec(o[12]),
                    _dec(o[13]),
                    _dec(
                        (o[8] or _ZERO)
                        - (o[10] or _ZERO)
                        - (o[11] or _ZERO)
                        - (o[12] or _ZERO)
                        - (o[13] or _ZERO)
                    ),
                ]
                for o in detail.orders
            ],
            bold,
        )
        _write_sheet(
            wb.create_sheet("Statements"),
            [
                "channel",
                "statement id",
                "period start",
                "period end",
                "gross sales",
                "total fees",
                "total vat",
                "net payable",
                "payout id",
                "invoice",
            ],
            [
                [
                    s.channel,
                    s.statement_id,
                    s.period_start,
                    s.period_end,
                    _dec(s.gross_sales),
                    _dec(s.total_fees),
                    _dec(s.total_vat),
                    _dec(s.net_payable),
                    s.payout_transfer_id or "",
                    "yes" if s.invoice_object_key else "",
                ]
                for s in detail.statements
            ],
            bold,
        )
        _write_sheet(
            wb.create_sheet("Statement Lines"),
            [
                "channel",
                "statement id",
                "order id",
                "date",
                "type",
                "fee category",
                "description",
                "amount",
            ],
            [
                [
                    ln.channel,
                    ln.statement_id,
                    ln.external_order_id or "",
                    ln.line_date,
                    ln.line_type or "",
                    ln.fee_category or "",
                    ln.description or "",
                    _dec(ln.amount),
                ]
                for ln in detail.statement_lines
            ],
            bold,
        )
        _write_sheet(
            wb.create_sheet("Payouts"),
            [
                "channel",
                "transfer id",
                "date",
                "amount",
                "status",
                "statement id",
                "reference",
            ],
            [
                [
                    p.channel,
                    p.transfer_id,
                    p.transfer_date,
                    _dec(p.transfer_amount),
                    p.transfer_status or "",
                    p.statement_id or "",
                    p.payment_reference or "",
                ]
                for p in detail.payouts
            ],
            bold,
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _order_channel_label(source: str | None, aggregator_channel: str | None) -> str:
    if source == "aggregator":
        return aggregator_channel or "aggregator"
    if source == "online":
        return "website"
    if source == "cashier":
        return "counter"
    return source or ""


def _body_html(label: str, detail: ReportDetail) -> str:
    return (
        f"<p>Daily sales for <strong>{label}</strong> is attached as a "
        f"spreadsheet.</p><p>Five tabs, all for this trading day: <strong>Summary"
        f"</strong> (branch×channel matrix with totals), <strong>Orders</strong> "
        f"({len(detail.orders)}), <strong>Statements</strong> "
        f"({len(detail.statements)}), <strong>Statement Lines</strong> "
        f"({len(detail.statement_lines)}) and <strong>Payouts</strong> "
        f"({len(detail.payouts)}). Delivered trade only.</p>"
    )


async def send(
    db: AsyncSession, *, date_from: str, date_to: str, recipients: list[str]
) -> dict:
    """Build the report and mail it, once per recipient, journalled."""
    report = await build(db, date_from=date_from, date_to=date_to)
    detail = await build_detail(db, date_from=date_from, date_to=date_to)
    xlsx = to_xlsx(report, detail)

    single = date_from == date_to
    label = date_from if single else f"{date_from} to {date_to}"
    subject = f"Melting Moments — Daily sales {label}"
    filename = f"daily-sales-{date_from}{'' if single else '_' + date_to}.xlsx"
    html = _body_html(label, detail)

    outcomes = []
    for recipient in recipients:
        result = await email_service.send_with_attachment(
            recipient,
            subject,
            html,
            filename=filename,
            content=xlsx,
            template=_TEMPLATE,
        )
        outcomes.append(
            {
                "recipient": recipient,
                "status": result["status"],
                "error": result.get("error"),
            }
        )
    return {"subject": subject, "rows": len(report.rows), "sent": outcomes}


# ── The daily loop ───────────────────────────────────────────────────────────


#: How many recently-ended business days a tick will still send if they were
#: missed. A day is sent once every branch's configured close for it, plus
#: `_CLOSE_BUFFER`, is in the past — so a deploy, a restart, or a slow tick around
#: closing time no longer loses the night: the next tick within this look-back
#: catches it up. Three days covers a weekend outage without ever backfilling
#: ancient history (the `_already_sent` guard skips anything already mailed).
_CATCHUP_DAYS = 3


async def _already_sent(db: AsyncSession, business_date: str) -> bool:
    """Whether a report for this date already went — the subject carries it."""
    count = (
        await db.execute(
            select(func.count(EmailLog.id)).where(
                EmailLog.template == _TEMPLATE,
                EmailLog.subject.like(f"%{business_date}%"),
                EmailLog.status == "sent",
            )
        )
    ).scalar_one()
    return count > 0


def _branch_close(branch: Branch, day: date) -> datetime | None:
    """The instant `branch` shuts on trading day `day`, or None if its configured
    hours do not parse.

    `opening_to` is a wall-clock ``"HH:MM"`` on the shop's own clock. A branch
    whose close is at or before its open trades past midnight (09:00 → 02:00), so
    its close for `day` lands on the following morning — the +1440 shift, the same
    reading `trading_hours` uses everywhere else. Returned as a real instant so
    the caller can compare it to `now` without any date arithmetic of its own."""
    opens = trading_hours.minutes_of(branch.opening_from)
    closes = trading_hours.minutes_of(branch.opening_to)
    if closes is None:
        return None
    minute = closes + 1440 if opens is not None and closes <= opens else closes
    return trading_hours.at_minute(day, minute)


def _last_close(branches: list[Branch], day: date) -> datetime | None:
    """When the *last* of these branches shuts on `day` — the moment the estate's
    trading for that day is over. None if no branch's hours parse."""
    closes = [c for b in branches if (c := _branch_close(b, day)) is not None]
    return max(closes) if closes else None


async def _tick(db: AsyncSession, now: datetime | None = None) -> None:
    tz = await business_day_service.resolve_timezone(db)
    now = now or datetime.now(tz)

    branches = (
        (await db.execute(select(Branch).where(Branch.is_active.is_(True))))
        .scalars()
        .all()
    )
    if not branches:
        return

    # Candidate business days: today back through the look-back, oldest first. A
    # day is due once the *last* branch's close for it, plus the settle buffer,
    # has passed. For a past-midnight branch that close is the following morning,
    # and because the comparison is on the instant that resolves itself — no
    # frontier, no rollover arithmetic here. Today is included so tonight's report
    # goes out tonight; it is simply held until its own close has passed.
    today = now.astimezone(trading_hours.TZ).date()
    pending: list[str] = []
    for offset in range(_CATCHUP_DAYS, -1, -1):
        day = today - timedelta(days=offset)
        last_close = _last_close(branches, day)
        if last_close is None or now < last_close + _CLOSE_BUFFER:
            continue
        business_date = day.isoformat()
        if not await _already_sent(db, business_date):
            pending.append(business_date)
    if not pending:
        return

    async with advisory_lock.held(
        _ADVISORY_LOCK_KEY, name="daily sales report"
    ) as mine:
        if not mine:
            return
        for business_date in pending:
            # Re-check under the lock: another worker may have sent it in the gap.
            if await _already_sent(db, business_date):
                continue
            summary = await send(
                db,
                date_from=business_date,
                date_to=business_date,
                recipients=DEFAULT_RECIPIENTS,
            )
            logger.info(
                "Daily sales report sent for %s: %s", business_date, summary["sent"]
            )


async def run_forever() -> None:
    """Check every ten minutes; send each completed business day once, catching
    up any recent day a restart or outage caused to be missed."""
    logger.info("Daily sales report loop started")
    while True:
        try:
            async with AsyncSessionFactory() as db:
                await _tick(db)
        except asyncio.CancelledError:
            raise
        except Exception:  # noqa: BLE001 — a bad tick must not kill the loop
            logger.exception("Daily sales report tick failed")
        await asyncio.sleep(_TICK_SECONDS)
